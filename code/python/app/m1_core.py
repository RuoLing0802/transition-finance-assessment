from __future__ import annotations

import hashlib
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .config import (
    CATALOG_RULE_VERSION,
    CATALOG_SHEET,
    ENERGY_SPECS,
    EXPECTED_SHEETS,
    FIELD_CONTRACT_VERSION,
    INPUT_SHEETS,
    OPERATING_SPECS,
    REFERENCE_SHEET,
    REQUIRED_HEADERS,
    RULE_VERSION,
    SIMULATED_DATA_NOTICE,
)
from .schemas import Issue


DYNAMIC_AUDIT_KEYS = {
    "batch_id",
    "generated_at",
    "received_at",
    "report_id",
    "report_path",
    "report_relative_path",
    "source_path",
}


def deterministic_projection(value: Any) -> Any:
    """Remove run-specific audit metadata while preserving business results."""
    if isinstance(value, dict):
        return {
            key: deterministic_projection(item)
            for key, item in sorted(value.items())
            if key not in DYNAMIC_AUDIT_KEYS
        }
    if isinstance(value, list):
        return [deterministic_projection(item) for item in value]
    return value


def jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"[\s\u3000、，,。；;：:（）()\[\]{}\-_/]", "", text)


def split_sources(value: Any) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in re.split(r"[、,，;；/、|]+", str(value)) if part.strip()]


def excel_column(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def cell_reference(row_number: int | None, field: str | None, headers: list[str]) -> str | None:
    if row_number is None or field is None or field not in headers:
        return None
    return f"{excel_column(headers.index(field) + 1)}{row_number}"


class WorkbookAnalyzer:
    """Read and validate one uploaded workbook without changing source values."""

    def __init__(self, source_path: Path, batch_id: str) -> None:
        self.source_path = source_path
        self.batch_id = batch_id
        self.headers: dict[str, list[str]] = {}
        self.rows: dict[str, list[dict[str, Any]]] = {}
        self.sheet_names: list[str] = []
        self.validation_issues: list[dict[str, Any]] = []
        self._loaded = False

    def load(self) -> None:
        workbook = load_workbook(self.source_path, read_only=True, data_only=True)
        try:
            self.sheet_names = list(workbook.sheetnames)
            for sheet_name in EXPECTED_SHEETS:
                if sheet_name not in workbook.sheetnames:
                    self.headers[sheet_name] = []
                    self.rows[sheet_name] = []
                    continue
                worksheet = workbook[sheet_name]
                iterator = worksheet.iter_rows(values_only=True)
                first_row = next(iterator, tuple())
                headers = [str(value).strip() if value is not None else "" for value in first_row]
                self.headers[sheet_name] = headers
                sheet_rows: list[dict[str, Any]] = []
                for row_number, values in enumerate(iterator, start=2):
                    if not any(value is not None for value in values):
                        continue
                    row: dict[str, Any] = {
                        header: jsonable(values[index])
                        for index, header in enumerate(headers)
                        if header
                        and index < len(values)
                    }
                    row["_row_number"] = row_number
                    sheet_rows.append(row)
                self.rows[sheet_name] = sheet_rows
        finally:
            workbook.close()
        self._loaded = True

    def _issue(
        self,
        *,
        issue_id: str,
        company_code: str | None,
        sheet_name: str | None,
        row_number: int | None,
        field: str | None,
        rule: str,
        severity: str,
        message: str,
        original_value: Any = None,
    ) -> None:
        headers = self.headers.get(sheet_name or "", [])
        self.validation_issues.append(
            Issue(
                issue_id=issue_id,
                batch_id=self.batch_id,
                company_code=company_code,
                sheet_name=sheet_name,
                row_number=row_number,
                field=field,
                rule=rule,
                severity=severity,
                message=message,
                original_value=jsonable(original_value),
                evidence={
                    "cell": cell_reference(row_number, field, headers),
                    "rule_version": RULE_VERSION,
                },
            ).as_dict()
        )

    def validate(self) -> dict[str, Any]:
        if not self._loaded:
            self.load()

        expected_set = set(EXPECTED_SHEETS)
        actual_set = set(self.sheet_names)
        for sheet_name in EXPECTED_SHEETS:
            if sheet_name not in actual_set:
                self._issue(
                    issue_id=f"structure-missing-sheet-{sheet_name}",
                    company_code=None,
                    sheet_name=sheet_name,
                    row_number=None,
                    field=None,
                    rule="required_sheet",
                    severity="error",
                    message=f"缺少必需工作表：{sheet_name}",
                )
        for sheet_name in sorted(actual_set - expected_set):
            self._issue(
                issue_id=f"structure-unexpected-sheet-{sheet_name}",
                company_code=None,
                sheet_name=sheet_name,
                row_number=None,
                field=None,
                rule="exact_five_sheets",
                severity="error",
                message=f"发现未纳入M1契约的额外工作表：{sheet_name}",
            )

        for sheet_name in EXPECTED_SHEETS:
            actual_headers = self.headers.get(sheet_name, [])
            expected_headers = REQUIRED_HEADERS[sheet_name]
            if actual_headers != expected_headers:
                self._issue(
                    issue_id=f"structure-header-{sheet_name}",
                    company_code=None,
                    sheet_name=sheet_name,
                    row_number=1,
                    field=None,
                    rule="required_headers",
                    severity="error",
                    message=f"{sheet_name}表头与M1契约不一致",
                    original_value={"actual": actual_headers, "expected": expected_headers},
                )
                continue
            self._validate_rows(sheet_name)

        self._validate_keys_and_associations()
        company_codes = sorted(
            {
                str(row.get("企业代号"))
                for row in self.rows.get("基本信息", [])
                if row.get("企业代号") not in (None, "")
            }
        )
        error_count = sum(issue["severity"] == "error" for issue in self.validation_issues)
        warning_count = sum(issue["severity"] == "warning" for issue in self.validation_issues)
        status = "failed" if error_count else "passed_with_warnings" if warning_count else "passed"
        return {
            "status": status,
            "rule_version": RULE_VERSION,
            "catalog_rule_version": CATALOG_RULE_VERSION,
            "field_contract_version": FIELD_CONTRACT_VERSION,
            "simulated_data": True,
            "data_notice": SIMULATED_DATA_NOTICE,
            "expected_sheets": EXPECTED_SHEETS,
            "actual_sheets": self.sheet_names,
            "sheet_roles": {
                "input": INPUT_SHEETS,
                "catalog": [CATALOG_SHEET],
                "reference_only": [REFERENCE_SHEET],
            },
            "sheet_summary": {
                sheet: {
                    "row_count": len(self.rows.get(sheet, [])),
                    "column_count": len(self.headers.get(sheet, [])),
                    "headers_match": self.headers.get(sheet, []) == REQUIRED_HEADERS[sheet],
                }
                for sheet in EXPECTED_SHEETS
            },
            "company_codes": company_codes,
            "issue_count": len(self.validation_issues),
            "error_count": error_count,
            "warning_count": warning_count,
            "validation_issues": self.validation_issues,
        }

    def _validate_rows(self, sheet_name: str) -> None:
        rows = self.rows.get(sheet_name, [])
        text_fields = set(REQUIRED_HEADERS[sheet_name]) - {"企业代号"}
        numeric_fields: set[str] = set()
        integer_fields: set[str] = set()
        boolean_fields: set[str] = set()
        ratio_fields: set[str] = set()
        if sheet_name == "基本信息":
            integer_fields = {"成立年份"}
            text_fields -= integer_fields
        elif sheet_name == "能耗信息":
            integer_fields = {"主用能项数量"}
            numeric_fields = {
                header
                for header in REQUIRED_HEADERS[sheet_name]
                if "消费量" in header or header in {"2024年主要产品产量", "2025年主要产品产量", "2024年营业收入（万元）", "2025年营业收入（万元）"}
            }
            text_fields -= numeric_fields | integer_fields
        elif sheet_name == "补充信息":
            integer_fields = {"核心设备平均投运年限"}
            numeric_fields = {"2025年绿电比例"}
            ratio_fields = {"2025年绿电比例"}
            boolean_fields = {
                "是否已建设能耗/碳排在线管理系统",
                "是否已开展余热/余压回收",
                "是否已建设分布式光伏",
            }
            text_fields -= numeric_fields | integer_fields | boolean_fields
        elif sheet_name == CATALOG_SHEET:
            text_fields = set(REQUIRED_HEADERS[sheet_name])
        elif sheet_name == REFERENCE_SHEET:
            text_fields = set(REQUIRED_HEADERS[sheet_name])

        for row in rows:
            row_number = row["_row_number"]
            company_code = row.get("企业代号")
            for field in text_fields:
                value = row.get(field)
                if value is not None and not isinstance(value, str):
                    self._issue(
                        issue_id=f"type-{sheet_name}-{row_number}-{field}",
                        company_code=company_code,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        field=field,
                        rule="text_type",
                        severity="error",
                        message="文本字段不是文本类型",
                        original_value=value,
                    )
            for field in integer_fields:
                value = row.get(field)
                if value is not None and (not isinstance(value, int) or isinstance(value, bool)):
                    self._issue(
                        issue_id=f"type-{sheet_name}-{row_number}-{field}",
                        company_code=company_code,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        field=field,
                        rule="integer_type",
                        severity="error",
                        message="整数型字段不是整数",
                        original_value=value,
                    )
                if isinstance(value, int) and value < 0:
                    self._issue(
                        issue_id=f"range-{sheet_name}-{row_number}-{field}",
                        company_code=company_code,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        field=field,
                        rule="non_negative",
                        severity="error",
                        message="整数型字段不能为负数",
                        original_value=value,
                    )
            for field in numeric_fields:
                value = row.get(field)
                if value is not None and (not isinstance(value, (int, float)) or isinstance(value, bool)):
                    self._issue(
                        issue_id=f"type-{sheet_name}-{row_number}-{field}",
                        company_code=company_code,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        field=field,
                        rule="numeric_type",
                        severity="error",
                        message="数值字段不是数值类型",
                        original_value=value,
                    )
                if isinstance(value, (int, float)) and value < 0:
                    self._issue(
                        issue_id=f"range-{sheet_name}-{row_number}-{field}",
                        company_code=company_code,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        field=field,
                        rule="non_negative",
                        severity="error",
                        message="数值字段不能为负数",
                        original_value=value,
                    )
            for field in boolean_fields:
                value = row.get(field)
                if value is not None and value not in {"是", "否", True, False}:
                    self._issue(
                        issue_id=f"value-{sheet_name}-{row_number}-{field}",
                        company_code=company_code,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        field=field,
                        rule="boolean_enum",
                        severity="error",
                        message="布尔字段只能为“是”或“否”",
                        original_value=value,
                    )
            for field in ratio_fields:
                value = row.get(field)
                if isinstance(value, (int, float)) and not 0 <= value <= 1:
                    self._issue(
                        issue_id=f"range-{sheet_name}-{row_number}-{field}",
                        company_code=company_code,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        field=field,
                        rule="ratio_range",
                        severity="error",
                        message="绿电比例必须位于0—1之间",
                        original_value=value,
                    )
            if sheet_name == "基本信息":
                value = row.get("成立年份")
                if isinstance(value, int) and not 1900 <= value <= 2026:
                    self._issue(
                        issue_id=f"range-{sheet_name}-{row_number}-成立年份",
                        company_code=company_code,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        field="成立年份",
                        rule="year_range",
                        severity="error",
                        message="成立年份不在M1检查范围内（1900—2026）",
                        original_value=value,
                    )

    def _validate_keys_and_associations(self) -> None:
        key_sets: dict[str, set[str]] = {}
        for sheet_name in [*INPUT_SHEETS, REFERENCE_SHEET]:
            rows = self.rows.get(sheet_name, [])
            current: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for row in rows:
                value = row.get("企业代号")
                if value in (None, ""):
                    self._issue(
                        issue_id=f"key-empty-{sheet_name}-{row['_row_number']}",
                        company_code=None,
                        sheet_name=sheet_name,
                        row_number=row["_row_number"],
                        field="企业代号",
                        rule="non_empty_key",
                        severity="error",
                        message="企业代号为空",
                        original_value=value,
                    )
                    continue
                if not isinstance(value, str):
                    self._issue(
                        issue_id=f"key-type-{sheet_name}-{row['_row_number']}",
                        company_code=None,
                        sheet_name=sheet_name,
                        row_number=row["_row_number"],
                        field="企业代号",
                        rule="text_key",
                        severity="error",
                        message="企业代号必须为文本",
                        original_value=value,
                    )
                current[str(value)].append(row)
            key_sets[sheet_name] = set(current)
            for code, duplicate_rows in current.items():
                if len(duplicate_rows) > 1:
                    self._issue(
                        issue_id=f"key-duplicate-{sheet_name}-{code}",
                        company_code=code,
                        sheet_name=sheet_name,
                        row_number=duplicate_rows[0]["_row_number"],
                        field="企业代号",
                        rule="unique_key",
                        severity="error",
                        message=f"企业代号在{sheet_name}表重复出现{len(duplicate_rows)}次",
                        original_value=code,
                    )
        baseline = key_sets.get("基本信息", set())
        for sheet_name in ["能耗信息", "补充信息", REFERENCE_SHEET]:
            missing = sorted(baseline - key_sets.get(sheet_name, set()))
            extra = sorted(key_sets.get(sheet_name, set()) - baseline)
            if missing or extra:
                self._issue(
                    issue_id=f"association-{sheet_name}",
                    company_code=None,
                    sheet_name=sheet_name,
                    row_number=None,
                    field="企业代号",
                    rule="cross_sheet_key_set",
                    severity="error",
                    message=f"与基本信息表企业代号集合不一致：缺失{len(missing)}个，额外{len(extra)}个",
                    original_value={"missing": missing[:20], "extra": extra[:20]},
                )

    def _rows_by_code(self, sheet_name: str) -> dict[str, dict[str, Any]]:
        return {
            str(row.get("企业代号")): row
            for row in self.rows.get(sheet_name, [])
            if row.get("企业代号") not in (None, "")
        }

    def _company_issue(self, issue: Issue) -> dict[str, Any]:
        return issue.as_dict()

    def _derived_issue(
        self,
        *,
        code: str,
        rule: str,
        severity: str,
        message: str,
        sheet_name: str,
        row_number: int | None,
        field: str | None,
        original_value: Any = None,
    ) -> dict[str, Any]:
        headers = self.headers.get(sheet_name, [])
        return Issue(
            issue_id=f"derived-{rule}-{code}-{field or 'row'}",
            batch_id=self.batch_id,
            company_code=code,
            sheet_name=sheet_name,
            row_number=row_number,
            field=field,
            rule=rule,
            severity=severity,
            message=message,
            original_value=jsonable(original_value),
            evidence={
                "cell": cell_reference(row_number, field, headers),
                "rule_version": RULE_VERSION,
                "provisional": True,
            },
        ).as_dict()

    @staticmethod
    def _clean_row(row: dict[str, Any] | None) -> dict[str, Any] | None:
        if row is None:
            return None
        return {key: value for key, value in row.items() if not key.startswith("_")}

    def energy_trend(self, code: str, energy_row: dict[str, Any], row_number: int | None) -> dict[str, Any]:
        resources: list[dict[str, Any]] = []
        for name, year_2024, year_2025, unit in ENERGY_SPECS:
            value_2024 = energy_row.get(year_2024)
            value_2025 = energy_row.get(year_2025)
            resources.append(
                self._trend_item(
                    name=name,
                    year_2024=value_2024,
                    year_2025=value_2025,
                    unit=unit,
                    status=self._trend_status(value_2024, value_2025),
                )
            )
        operating: list[dict[str, Any]] = []
        for name, year_2024, year_2025, unit in OPERATING_SPECS:
            value_2024 = energy_row.get(year_2024)
            value_2025 = energy_row.get(year_2025)
            operating.append(
                self._trend_item(
                    name=name,
                    year_2024=value_2024,
                    year_2025=value_2025,
                    unit=unit or energy_row.get("产量单位"),
                    status=self._trend_status(value_2024, value_2025),
                )
            )
        return {
            "company_code": code,
            "years": [2024, 2025],
            "resources": resources,
            "operating_metrics": operating,
            "resource_names_present": [
                name
                for name, year_2024, year_2025, _unit in ENERGY_SPECS
                if energy_row.get(year_2024) is not None or energy_row.get(year_2025) is not None
            ],
            "production_unit": energy_row.get("产量单位"),
            "calculation_note": "仅比较工作簿原始值；未取得组织/核算边界、折标系数和排放因子，不输出正式碳排放量。",
        }

    @staticmethod
    def _trend_status(value_2024: Any, value_2025: Any) -> str:
        if value_2024 is None and value_2025 is None:
            return "成对缺失"
        if value_2024 is None or value_2025 is None:
            return "年度单边缺失"
        return "可比较"

    @staticmethod
    def _trend_item(*, name: str, year_2024: Any, year_2025: Any, unit: Any, status: str) -> dict[str, Any]:
        delta = None
        rate = None
        if status == "可比较":
            if not all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in (year_2024, year_2025)):
                status = "类型异常"
            else:
                delta = year_2025 - year_2024
                if year_2024 != 0:
                    rate = delta / year_2024
        return {
            "name": name,
            "unit": unit,
            "2024": year_2024,
            "2025": year_2025,
            "change": delta,
            "change_rate": rate,
            "status": status,
        }

    def company_quality_issues(
        self,
        code: str,
        basic_row: dict[str, Any],
        energy_row: dict[str, Any],
        supplement_row: dict[str, Any],
    ) -> list[dict[str, Any]]:
        issues = [
            issue
            for issue in self.validation_issues
            if issue.get("company_code") in (None, code)
        ]
        row_number = energy_row.get("_row_number")
        for name, year_2024, year_2025, _unit in ENERGY_SPECS:
            value_2024 = energy_row.get(year_2024)
            value_2025 = energy_row.get(year_2025)
            if value_2024 is None and value_2025 is None:
                issues.append(
                    self._derived_issue(
                        code=code,
                        rule="annual_pair_missing",
                        severity="warning",
                        message=f"{name} 2024和2025年均缺失，未按零处理",
                        sheet_name="能耗信息",
                        row_number=row_number,
                        field=year_2024,
                        original_value={"2024": value_2024, "2025": value_2025},
                    )
                )
            elif value_2024 is None or value_2025 is None:
                issues.append(
                    self._derived_issue(
                        code=code,
                        rule="annual_one_sided_missing",
                        severity="warning",
                        message=f"{name}存在年度单边缺失，不能计算完整变化率",
                        sheet_name="能耗信息",
                        row_number=row_number,
                        field=year_2024 if value_2024 is None else year_2025,
                        original_value={"2024": value_2024, "2025": value_2025},
                    )
                )

        production_2024 = energy_row.get("2024年主要产品产量")
        production_2025 = energy_row.get("2025年主要产品产量")
        production_unit = energy_row.get("产量单位")
        production_missing = production_2024 is None and production_2025 is None
        if production_missing and production_unit is None:
            issues.append(
                self._derived_issue(
                    code=code,
                    rule="production_unit_pair_missing",
                    severity="warning",
                    message="主要产品产量和产量单位成对缺失，未按零处理",
                    sheet_name="能耗信息",
                    row_number=row_number,
                    field="产量单位",
                )
            )
        elif production_missing != (production_unit is None):
            issues.append(
                self._derived_issue(
                    code=code,
                    rule="production_unit_pair_mismatch",
                    severity="warning",
                    message="主要产品产量与产量单位一边缺失，需人工复核",
                    sheet_name="能耗信息",
                    row_number=row_number,
                    field="产量单位",
                    original_value={
                        "2024年主要产品产量": production_2024,
                        "2025年主要产品产量": production_2025,
                        "产量单位": production_unit,
                    },
                )
            )

        declared_count = energy_row.get("主用能项数量")
        if isinstance(declared_count, int):
            observed_count = sum(
                1
                for _name, year_2024, year_2025, _unit in ENERGY_SPECS
                if energy_row.get(year_2024) is not None or energy_row.get(year_2025) is not None
            )
            text_count = len(split_sources(energy_row.get("主要用能来源")))
            if observed_count != declared_count or text_count != declared_count:
                issues.append(
                    self._derived_issue(
                        code=code,
                        rule="energy_source_count_soft_check",
                        severity="warning",
                        message="主要用能来源文本、主用能项数量与数值能源项数量不完全一致，当前仅作软校验",
                        sheet_name="能耗信息",
                        row_number=row_number,
                        field="主用能项数量",
                        original_value={
                            "declared_count": declared_count,
                            "text_count": text_count,
                            "observed_numeric_count": observed_count,
                        },
                    )
                )

        founded_year = basic_row.get("成立年份")
        equipment_age = supplement_row.get("核心设备平均投运年限")
        if isinstance(founded_year, int) and isinstance(equipment_age, int):
            operating_years = 2025 - founded_year
            if equipment_age > operating_years:
                issues.append(
                    self._derived_issue(
                        code=code,
                        rule="equipment_age_warning",
                        severity="warning",
                        message="核心设备平均投运年限超过企业自成立至2025年的年数，需确认期间口径",
                        sheet_name="补充信息",
                        row_number=supplement_row.get("_row_number"),
                        field="核心设备平均投运年限",
                        original_value={"equipment_age": equipment_age, "operating_years": operating_years},
                    )
                )
        return sorted(issues, key=lambda item: (item.get("severity", ""), item.get("issue_id", "")))

    def catalog_matches(self, code: str, basic_row: dict[str, Any], supplement_row: dict[str, Any]) -> dict[str, Any]:
        industry = basic_row.get("行业")
        subindustry = basic_row.get("细分行业/领域")
        product = supplement_row.get("主要产品/服务")
        industry_norm = normalize_text(industry)
        subindustry_norm = normalize_text(subindustry)
        product_norm = normalize_text(product)
        candidates: list[dict[str, Any]] = []
        for row in self.rows.get(CATALOG_SHEET, []):
            if normalize_text(row.get("行业")) != industry_norm or not industry_norm:
                continue
            category = row.get("类别/领域")
            path = row.get("转型路径")
            description = row.get("说明")
            category_norm = normalize_text(category)
            path_norm = normalize_text(path)
            description_norm = normalize_text(description)
            provisional_sort_key = 5
            reasons = ["行业完全匹配"]
            if not category_norm:
                provisional_sort_key += 1
                reasons.append("目录类别/领域为空，作为行业级候选")
            elif category_norm == subindustry_norm:
                provisional_sort_key += 4
                reasons.append("细分行业/领域完全匹配")
            elif category_norm in subindustry_norm or subindustry_norm in category_norm:
                provisional_sort_key += 3
                reasons.append("细分行业/领域包含匹配")
            if product_norm and category_norm and (category_norm in product_norm or product_norm in category_norm):
                provisional_sort_key += 3
                reasons.append("主要产品/服务与目录类别有包含匹配")
            elif product_norm and any(token and token in f"{path_norm}{description_norm}" for token in [product_norm]):
                provisional_sort_key += 1
                reasons.append("主要产品/服务在路径或说明中出现")
            candidates.append(
                {
                    "catalog_row_id": f"转型目录!A{row['_row_number']}",
                    "row_number": row["_row_number"],
                    "industry": row.get("行业"),
                    "category": category,
                    "transition_path": path,
                    "description": description,
                    "provisional_sort_key": provisional_sort_key,
                    "provisional_rule": True,
                    "match_reasons": reasons,
                }
            )
        candidates.sort(key=lambda item: (-item["provisional_sort_key"], item["row_number"]))
        top_key = candidates[0]["provisional_sort_key"] if candidates else None
        top_count = sum(item["provisional_sort_key"] == top_key for item in candidates) if top_key is not None else 0
        missing_detail = not subindustry_norm and not product_norm
        if not candidates:
            status = "覆盖缺口"
        elif missing_detail:
            status = "缺少细分行业/产品信息"
        else:
            status = "存在歧义" if top_count > 1 else "候选已生成"
        for candidate in candidates:
            candidate["top_provisional_candidate"] = candidate["provisional_sort_key"] == top_key
            candidate["ambiguity"] = top_count > 1 and candidate["provisional_sort_key"] == top_key
        return {
            "company_code": code,
            "catalog_rule_version": CATALOG_RULE_VERSION,
            "status": status,
            "query": {"industry": industry, "subindustry": subindustry, "product_or_service": product},
            "match_policy": "行业必需匹配；细分行业/领域和产品/服务用于暂定排序；该排序值不是评分，不能代表企业优劣；路径编号不作为唯一键",
            "candidates": candidates[:20],
            "total_industry_candidates": len(candidates),
            "manual_review_required": status != "候选已生成",
        }

    def reference_comparison(
        self,
        code: str,
        input_data: dict[str, Any],
        energy_trend: dict[str, Any],
        catalog: dict[str, Any],
        issues: list[dict[str, Any]],
    ) -> dict[str, Any]:
        reference_row = self._rows_by_code(REFERENCE_SHEET).get(code)
        reference_fields = [field for field in REQUIRED_HEADERS[REFERENCE_SHEET] if field != "企业代号"]
        input_field_names = {
            field
            for table in input_data.values()
            if isinstance(table, dict)
            for field in table
        }
        excluded_from_input = [field for field in reference_fields if field not in input_field_names]
        leakage_passed = set(reference_fields).isdisjoint(input_field_names)
        if reference_row is None:
            return {
                "status": "未找到参考结论",
                "reference_fields": {},
                "reference_fields_present": 0,
                "field_presence_comparison": [],
                "comparison_items": self._comparison_items(
                    reference={}, energy_trend=energy_trend, catalog=catalog, reference_available=False
                ),
                "comparison_summary": {"参考值缺失": 5},
                "leakage_guard": {
                    "status": "passed" if leakage_passed else "failed",
                    "reference_fields_excluded_from_input": excluded_from_input,
                },
                "flow_summary": self._flow_summary(energy_trend, catalog, issues),
                "comparison_notice": "参考结论层独立加载；本M1不将其作为模型输入、特征或标签。",
            }
        reference = self._clean_row(reference_row) or {}
        field_presence = [
            {
                "field": field,
                "reference_value_present": reference.get(field) not in (None, ""),
                "flow_source": "独立输入与规则/质量流程",
                "comparison_status": "仅作人工对照，不构成模型效果指标",
            }
            for field in reference_fields
        ]
        comparison_items = self._comparison_items(
            reference=reference, energy_trend=energy_trend, catalog=catalog, reference_available=True
        )
        return {
            "status": "已加载独立参考结论",
            "reference_fields": reference,
            "reference_fields_present": sum(item["reference_value_present"] for item in field_presence),
            "field_presence_comparison": field_presence,
            "comparison_items": comparison_items,
            "comparison_summary": {
                status: sum(item["status"] == status for item in comparison_items)
                for status in sorted({item["status"] for item in comparison_items})
            },
            "leakage_guard": {
                "status": "passed" if leakage_passed else "failed",
                "reference_fields_excluded_from_input": excluded_from_input,
                "reference_sheet": REFERENCE_SHEET,
                "rule": "reference_only_layer",
            },
            "flow_summary": self._flow_summary(energy_trend, catalog, issues),
            "comparison_notice": "该对照仅用于流程回放和人工审阅，不计算准确率、相关性或业务效果。",
        }

    @staticmethod
    def _comparison_items(
        *, reference: dict[str, Any], energy_trend: dict[str, Any], catalog: dict[str, Any], reference_available: bool
    ) -> list[dict[str, Any]]:
        """Create explicit, human-readable comparison mappings without scoring."""
        if not reference_available:
            missing_status = "参考值缺失"
            return [
                {
                    "mapping": "转型规划结论.主要用能特征 ↔ 能耗信息.主要用能来源/主用能项数量",
                    "reference_field": "主要用能特征",
                    "flow_source": "能耗信息",
                    "status": missing_status,
                    "reason": "当前企业没有对应的参考结论行，不能形成语义对照。",
                },
                {
                    "mapping": "转型规划结论.能耗数据关联要点 ↔ 能耗信息.2024—2025变化摘要",
                    "reference_field": "能耗数据关联要点",
                    "flow_source": "能耗信息/能耗变化分析",
                    "status": missing_status,
                    "reason": "当前企业没有对应的参考结论行，不能形成语义对照。",
                },
                {
                    "mapping": "转型规划结论.匹配的转型路径名称 ↔ 转型目录候选",
                    "reference_field": "匹配的转型路径名称",
                    "flow_source": "转型目录",
                    "status": missing_status,
                    "reason": "当前企业没有对应的参考结论行，不能形成语义对照。",
                },
                {
                    "mapping": "转型规划结论.建议改进方向 ↔ 质量问题与缺失提示",
                    "reference_field": "建议改进方向",
                    "flow_source": "质量校验",
                    "status": missing_status,
                    "reason": "当前企业没有对应的参考结论行，不能形成语义对照。",
                },
                {
                    "mapping": "转型规划结论.行动建议 ↔ M1流程输出",
                    "reference_field": "近/中/长期转型行动建议",
                    "flow_source": "M1流程输出",
                    "status": missing_status,
                    "reason": "M1不生成行动建议，不对缺失参考值做补写。",
                },
            ]

        def present(field: str) -> bool:
            return reference.get(field) not in (None, "")

        energy_sources = energy_trend.get("resources", [])
        comparable = sum(item.get("status") == "可比较" for item in energy_sources)
        reference_energy = normalize_text(reference.get("主要用能特征"))
        current_energy = normalize_text("、".join(energy_trend.get("resource_names_present", [])))
        if not present("主要用能特征"):
            energy_status, energy_reason = "参考值为空", "参考字段没有值，无法进行字面一致性观察。"
        elif reference_energy and current_energy and (reference_energy in current_energy or current_energy in reference_energy):
            energy_status, energy_reason = "发现字面重合", "参考用能特征与当前输入的能源项存在字面重合；这不是模型正确性结论。"
        else:
            energy_status, energy_reason = "未发现字面重合", "当前仅做字面观察；未重合不代表参考结论错误。"

        reference_trend = normalize_text(reference.get("能耗数据关联要点"))
        if not present("能耗数据关联要点"):
            trend_status, trend_reason = "参考值为空", "参考字段没有值，无法进行对照。"
        elif reference_trend and comparable:
            trend_status, trend_reason = "流程摘要可对照", f"当前流程提供{comparable}项年度可比较指标；不评价参考结论是否正确。"
        else:
            trend_status, trend_reason = "流程摘要不足", "当前流程没有足够的年度成对数值，不能形成完整变化摘要对照。"

        paths = [normalize_text(item.get("transition_path")) for item in catalog.get("candidates", [])]
        reference_path = normalize_text(reference.get("匹配的转型路径名称"))
        if not present("匹配的转型路径名称"):
            path_status, path_reason = "参考值为空", "参考字段没有值，无法进行对照。"
        elif reference_path and any(reference_path in path or path in reference_path for path in paths if path):
            path_status, path_reason = "发现字面重合", "参考路径与目录候选存在字面重合；候选仍需人工复核。"
        elif catalog.get("candidates"):
            path_status, path_reason = "未发现字面重合", "目录已生成候选但未发现字面重合；不代表参考结论错误。"
        else:
            path_status, path_reason = "候选不足", "当前没有目录候选，不能形成路径对照。"

        return [
            {
                "mapping": "转型规划结论.主要用能特征 ↔ 能耗信息.主要用能来源/主用能项数量",
                "reference_field": "主要用能特征",
                "flow_source": "能耗信息",
                "status": energy_status,
                "reason": energy_reason,
            },
            {
                "mapping": "转型规划结论.能耗数据关联要点 ↔ 能耗信息.2024—2025变化摘要",
                "reference_field": "能耗数据关联要点",
                "flow_source": "能耗信息/能耗变化分析",
                "status": trend_status,
                "reason": trend_reason,
            },
            {
                "mapping": "转型规划结论.匹配的转型路径名称 ↔ 转型目录候选",
                "reference_field": "匹配的转型路径名称",
                "flow_source": "转型目录",
                "status": path_status,
                "reason": path_reason,
            },
            {
                "mapping": "转型规划结论.建议改进方向 ↔ 质量问题与缺失提示",
                "reference_field": "建议改进方向",
                "flow_source": "质量校验",
                "status": "仅供人工对照",
                "reason": "M1不从参考结论反推出建议，仅展示质量问题和缺失提示供人工审阅。",
            },
            {
                "mapping": "转型规划结论.近/中/长期转型行动建议 ↔ M1流程输出",
                "reference_field": "近/中/长期转型行动建议",
                "flow_source": "M1流程输出",
                "status": "M1不生成",
                "reason": "行动建议仍属于参考输出层，M1不把它作为输入，也不自行生成替代结论。",
            },
        ]

    @staticmethod
    def _flow_summary(energy_trend: dict[str, Any], catalog: dict[str, Any], issues: list[dict[str, Any]]) -> dict[str, Any]:
        return {
            "quality_issue_count": len(issues),
            "quality_issue_by_severity": {
                severity: sum(issue.get("severity") == severity for issue in issues)
                for severity in ["error", "warning", "info"]
            },
            "energy_resource_count": len(energy_trend.get("resources", [])),
            "energy_comparable_count": sum(item.get("status") == "可比较" for item in energy_trend.get("resources", [])),
            "catalog_candidate_count": len(catalog.get("candidates", [])),
            "catalog_status": catalog.get("status"),
        }

    def analyze_company(self, code: str) -> dict[str, Any]:
        if not self._loaded:
            self.load()
        basic_row = self._rows_by_code("基本信息").get(code)
        energy_row = self._rows_by_code("能耗信息").get(code)
        supplement_row = self._rows_by_code("补充信息").get(code)
        if basic_row is None or energy_row is None or supplement_row is None:
            raise KeyError(f"企业代号不存在或输入表未闭合: {code}")
        input_data = {
            "basic_info": self._clean_row(basic_row),
            "energy_info": self._clean_row(energy_row),
            "supplement_info": self._clean_row(supplement_row),
        }
        issues = self.company_quality_issues(code, basic_row, energy_row, supplement_row)
        trend = self.energy_trend(code, energy_row, energy_row.get("_row_number"))
        catalog = self.catalog_matches(code, basic_row, supplement_row)
        comparison = self.reference_comparison(code, input_data, trend, catalog, issues)
        return {
            "batch_id": self.batch_id,
            "company_code": code,
            "data_status": "命题方脱敏模拟数据",
            "simulated_data": True,
            "data_notice": SIMULATED_DATA_NOTICE,
            "rule_version": RULE_VERSION,
            "field_contract_version": FIELD_CONTRACT_VERSION,
            "input_data": input_data,
            "quality_policy": "M1暂用错误/警告/提示分级；不作为最终评分、授信或业务阻断规则",
            "quality_issues": issues,
            "energy_trend": trend,
            "catalog_matches": catalog,
            "reference_comparison": comparison,
            "boundaries": {
                "reference_sheet_excluded_from_input": True,
                "catalog_is_rule_knowledge_layer": True,
                "formal_carbon_calculation": False,
                "formal_scoring": False,
                "lending_decision": False,
            },
        }

    def aggregate_company_quality(self, validation: dict[str, Any]) -> dict[str, Any]:
        """Aggregate derived company checks separately from structural validation."""
        if validation.get("status") == "failed":
            return {
                "status": "blocked",
                "issue_count": 0,
                "affected_company_count": 0,
                "by_severity": {"error": 0, "warning": 0, "info": 0},
                "by_rule": {},
                "notice": "批次存在结构或字段错误，未执行企业级派生质量聚合；详情和报告流程已阻断。",
            }
        basic_rows = self._rows_by_code("基本信息")
        energy_rows = self._rows_by_code("能耗信息")
        supplement_rows = self._rows_by_code("补充信息")
        derived_issues: list[dict[str, Any]] = []
        for code in sorted(validation.get("company_codes", [])):
            if code not in energy_rows or code not in supplement_rows:
                continue
            company_issues = self.company_quality_issues(
                code, basic_rows.get(code, {}), energy_rows[code], supplement_rows[code]
            )
            derived_issues.extend(issue for issue in company_issues if issue.get("issue_id", "").startswith("derived-"))
        by_rule: dict[str, int] = defaultdict(int)
        for issue in derived_issues:
            by_rule[issue.get("rule", "unknown")] += 1
        return {
            "status": "available",
            "issue_count": len(derived_issues),
            "affected_company_count": len({issue.get("company_code") for issue in derived_issues}),
            "by_severity": {
                severity: sum(issue.get("severity") == severity for issue in derived_issues)
                for severity in ["error", "warning", "info"]
            },
            "by_rule": dict(sorted(by_rule.items())),
            "sample_issues": sorted(
                derived_issues,
                key=lambda issue: (str(issue.get("company_code")), str(issue.get("rule")), str(issue.get("field"))),
            )[:20],
            "notice": "企业级派生质量问题来自输入值关系和年度成对检查；严重度为M1暂定，不是正式评分规则。",
        }

    def batch_summary(self, source_metadata: dict[str, Any]) -> dict[str, Any]:
        validation = self.validate()
        structural_issues = [
            issue for issue in validation["validation_issues"] if not issue.get("issue_id", "").startswith("derived-")
        ]
        quality_overview = {
            "structural": {
                "issue_count": len(structural_issues),
                "error_count": sum(issue.get("severity") == "error" for issue in structural_issues),
                "warning_count": sum(issue.get("severity") == "warning" for issue in structural_issues),
                "notice": "结构/字段校验与企业级派生质量问题分层展示。",
            },
            "enterprise_derived": self.aggregate_company_quality(validation),
        }
        validation["quality_overview"] = quality_overview
        return {
            "batch_id": self.batch_id,
            "source_filename": source_metadata["source_filename"],
            "file_size": source_metadata["file_size"],
            "sha256": source_metadata["sha256"],
            "received_at": source_metadata["received_at"],
            "simulated_data": True,
            "data_notice": SIMULATED_DATA_NOTICE,
            "status": validation["status"],
            "processing_stage": "completed",
            "rule_version": RULE_VERSION,
            "field_contract_version": FIELD_CONTRACT_VERSION,
            "quality_overview": quality_overview,
            "validation": validation,
        }


def workbook_fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
