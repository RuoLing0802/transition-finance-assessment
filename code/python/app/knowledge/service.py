from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

from ..config import PROJECT_ROOT, SIMULATED_DATA_NOTICE
from ..domain_store import DomainConflictError, DomainNotFoundError, DomainStore
from .policy import (
    ALLOWLIST_VERSION,
    BM25_WEIGHTS,
    BLOCKED_SOURCE_IDS,
    DECISION_DATE,
    INDUSTRIES,
    METADATA_ONLY_SOURCE_IDS,
    RANKING_CONFIG_VERSION,
    SEARCHABLE_SOURCE_IDS,
    TOKENIZER_VERSION,
    INDUSTRY_MAPPING_VERSION,
    has_reference_marker,
    industry_scope_for_source,
    most_restrictive,
    normalize_industry,
    normalize_query_tokens,
    normalize_search_text,
    public_url,
    RESEARCH_SCOPE_BASIS,
    RESEARCH_INDUSTRY_SCOPE,
    source_role_priority,
    visibility_rank,
)
from .schemas import KnowledgeSearchResponse, KnowledgeSearchResult


class KnowledgeBuildBlocked(DomainConflictError):
    """The manifest failed a hard M5 admission or leakage gate."""

    def __init__(self, message: str, report: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.report = report or {}


class KnowledgeIndexNotReady(DomainConflictError):
    """Normal retrieval cannot build or repair a global index."""

    code = "knowledge_index_not_ready"


@dataclass(frozen=True)
class KnowledgeAssetPaths:
    source_ledger: Path
    candidates: Path
    gold_tests: Path
    literature: Path
    intake_ledger: Path

    @classmethod
    def from_project_root(cls, project_root: Path = PROJECT_ROOT) -> "KnowledgeAssetPaths":
        root = project_root / "团队成果" / "03_数据政策评分"
        return cls(
            source_ledger=root / "08_国内外知识源抓取与核验台账.xlsx",
            candidates=root / "12_M5候选语料与切片清单.xlsx",
            gold_tests=root / "13_M5检索金标准与负向测试.xlsx",
            literature=root / "14_中文与国际学术文献证据表.xlsx",
            intake_ledger=root / "20_补充政策文献与国家因子库接收台账.xlsx",
        )


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sha256(value: bytes | str) -> str:
    payload = value.encode("utf-8") if isinstance(value, str) else value
    return hashlib.sha256(payload).hexdigest()


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _slug(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("_") or "item"


def _short_text(value: Any, limit: int = 2400) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:limit]


def _load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        rows = list(workbook[sheet_name].values)
    finally:
        workbook.close()
    if not rows:
        return []
    header = [str(item).strip() if item is not None else "" for item in rows[0]]
    return [
        {key: value for key, value in zip(header, row) if key}
        for row in rows[1:]
        if any(value not in (None, "") for value in row)
    ]


def _safe_asset_path(project_root: Path, raw_path: Any) -> Path | None:
    if raw_path in (None, ""):
        return None
    value = Path(str(raw_path).strip())
    if value.is_absolute() or ".." in value.parts:
        return None
    candidate = (project_root / value).resolve()
    if project_root.resolve() not in candidate.parents and candidate != project_root.resolve():
        return None
    return candidate if candidate.is_file() else None


def _literature_pdf_path(project_root: Path, availability: Any) -> Path | None:
    """Resolve the controlled local PDF path embedded in the literature ledger.

    The ledger deliberately keeps the human-readable receipt, path and hash in
    one cell.  Only the path portion is used for local extraction; the recorded
    SHA-256 remains the identity check and is never inferred from the title.
    """
    value = str(availability or "")
    match = re.search(r"；([^；]+\.pdf)；SHA-256=", value, flags=re.IGNORECASE)
    raw_path = match.group(1).strip() if match else ""
    path = _safe_asset_path(project_root, raw_path)
    return path


def _record_date_uncertain(*values: Any) -> bool:
    """Mark year-only dates without pretending they are a specific day."""
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        if re.fullmatch(r"\d{4}", text) or (re.search(r"(?:19|20)\d{2}", text) and not re.search(r"(?:19|20)\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", text)):
            return True
    return False


def _parse_full_date(value: Any) -> date | None:
    text = str(value or "").strip()
    if not text or re.fullmatch(r"\d{4}", text):
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _normalize_pdf_text(value: str) -> str:
    return re.sub(r"[ \t\r\f\v]+", " ", value.replace("\u00a0", " ")).strip()


def _read_pdf_pages(path: Path) -> list[str]:
    try:
        import fitz

        document = fitz.open(path)
        try:
            return [_normalize_pdf_text(page.get_text("text")) for page in document]
        finally:
            document.close()
    except Exception:
        return []


def _split_page(text: str, limit: int = 1200, overlap: int = 80) -> list[str]:
    if len(text) <= limit:
        return [text] if text else []
    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = min(len(text), start + limit)
        chunks.append(text[start:end])
        if end == len(text):
            break
        start = max(start + 1, end - overlap)
    return chunks


def _extract_research_pages(pages: list[str]) -> list[tuple[int, str, str]]:
    """Select deterministic abstract/conclusion pages; never index research full text."""
    selected: list[tuple[int, str, str]] = []
    for index, text in enumerate(pages, start=1):
        if not text:
            continue
        label = None
        if re.search(r"(?:^|\s)(摘要|关键词|研究结论|结论与启示|结论|结语)(?:\s|$|：|:)", text):
            label = "摘要/结论候选片段"
        if label:
            selected.append((index, label, text[:2400]))
    if not selected and pages:
        nonempty = [(index, text) for index, text in enumerate(pages, start=1) if text]
        if nonempty:
            selected = [(nonempty[0][0], "全文首个可读片段", nonempty[0][1][:1800])]
            if len(nonempty) > 1:
                selected.append((nonempty[-1][0], "全文末个可读片段", nonempty[-1][1][:1800]))
    # Keep the manifest compact and stable when a PDF repeats headings.
    return selected[:4]


class KnowledgeService:
    """M5 registry, index builder and run-scoped deterministic retrieval service."""

    def __init__(
        self,
        domain_store: DomainStore,
        *,
        analysis_loader: Callable[[str], tuple[dict[str, Any], dict[str, Any]]] | None = None,
        assets: KnowledgeAssetPaths | None = None,
        project_root: Path = PROJECT_ROOT,
    ) -> None:
        self.domain_store = domain_store
        self.analysis_loader = analysis_loader
        self.assets = assets or KnowledgeAssetPaths.from_project_root(project_root)
        self.project_root = project_root.resolve()
        self._fts_available: bool | None = None

    @property
    def fts_available(self) -> bool:
        if self._fts_available is not None:
            return self._fts_available
        try:
            with self.domain_store._connection() as connection:
                connection.execute("CREATE VIRTUAL TABLE IF NOT EXISTS temp.m5_fts_probe USING fts5(value)")
                connection.execute("DROP TABLE IF EXISTS temp.m5_fts_probe")
            self._fts_available = True
        except sqlite3.OperationalError:
            self._fts_available = False
        return self._fts_available

    def _source_rows(self) -> dict[str, dict[str, Any]]:
        result: dict[str, dict[str, Any]] = {}
        for sheet in ("01_来源台账", "02_新增来源"):
            for row in _load_rows(self.assets.source_ledger, sheet):
                source_id = str(row.get("source_id") or "").strip()
                if source_id:
                    result[source_id] = {**result.get(source_id, {}), **row, "source_record_file": f"08_国内外知识源抓取与核验台账.xlsx:{sheet}"}
        return result

    def _asset_rows(self) -> list[dict[str, Any]]:
        rows = _load_rows(self.assets.intake_ledger, "01_政策标准PDF")
        for row in rows:
            path = _safe_asset_path(self.project_root, row.get("首选相对路径"))
            row["_path"] = path
            row["_relative_path"] = path.relative_to(self.project_root).as_posix() if path else None
            expected_hash = str(row.get("SHA-256") or "").strip().lower()
            actual_hash = _sha256(path.read_bytes()) if path and path.is_file() else None
            row["_expected_sha256"] = expected_hash or None
            row["_actual_sha256"] = actual_hash
            row["_hash_status"] = (
                "missing" if not expected_hash else
                "unreadable" if actual_hash is None else
                "match" if actual_hash == expected_hash else "mismatch"
            )
        return rows

    @staticmethod
    def _asset_source_id(row: dict[str, Any]) -> str:
        linked = str(row.get("linked_source_id") or "").strip()
        if linked and linked != "待建立候选source_id":
            return linked
        return f"ASSET-{row.get('asset_id', 'UNKNOWN')}"

    def _literature_rows(self, sheet: str) -> list[dict[str, Any]]:
        return _load_rows(self.assets.literature, sheet)

    def _decision_for_source(
        self,
        source: dict[str, Any],
        asset: dict[str, Any] | None,
        *,
        mapping_method: str,
    ) -> dict[str, Any]:
        source_id = str(source.get("source_id") or "")
        admission = str(source.get("admission_status") or "")
        text_status = str((asset or {}).get("文本状态") or "")
        receive_status = str((asset or {}).get("接收结论") or "")
        path = (asset or {}).get("_path")
        source_verified = str(source.get("source_identity_status") or "") == "已核验"
        hash_status = str((asset or {}).get("_hash_status") or "missing")
        hash_ok = hash_status == "match"
        file_readable = bool(path and path.is_file() and "可检索文本" in text_status and hash_ok)
        has_official_url = bool(source.get("official_url"))
        has_hash = bool((asset or {}).get("SHA-256"))
        visibility = "metadata_only"
        basis = "未形成M5初始正文来源白名单；只登记元数据并保留人工复核边界"
        if source_id in BLOCKED_SOURCE_IDS or "不得正式入库" in admission:
            visibility = "blocked"
            basis = "来源台账或M5边界明确阻断"
        elif source_id in SEARCHABLE_SOURCE_IDS:
            if source_id == "MEE-GHG-QA-2025":
                visibility = "searchable_candidate"
                basis = "08来源台账已核验官方网页；M5仅提供解释性候选证据，不生成因子值"
            elif source_verified and has_official_url and has_hash and file_readable and path:
                visibility = "searchable_candidate"
                basis = "08来源身份、20号受控原件哈希、可读文本和页面定位共同满足M5候选门槛"
            elif hash_status in {"mismatch", "missing", "unreadable"}:
                visibility = "blocked"
                basis = f"受控原件SHA-256{hash_status}；禁止读取或生成正文证据"
            elif "0页" in receive_status or "不可解析" in text_status:
                visibility = "blocked"
                basis = "原件0页或不可解析，不能生成正文证据"
            else:
                visibility = "metadata_only"
                basis = "来源候选已登记，但正文可读性、哈希或定位仍不足"
        elif source_id in METADATA_ONLY_SOURCE_IDS:
            visibility = "metadata_only"
            basis = "M5提示词明确限定为来源元数据；正文待OCR/重新取得/定位"
        elif "0页" in receive_status or "不可解析" in text_status:
            visibility = "blocked"
            basis = "0页或损坏原件阻断正文"
        elif asset and (not source_verified or not has_official_url or not has_hash):
            visibility = "metadata_only"
            basis = "补充原件与候选来源不能以标题相似自动合并，需人工映射"
        elif source.get("role") == "factor_candidate":
            visibility = "blocked"
            basis = "具体排放因子仍属M6候选，M5禁止自动调用"
        elif asset:
            visibility = "metadata_only"
            basis = "未列入M5初始正文白名单的补充PDF，只登记来源元数据"
        if path and not path.is_file() and visibility == "searchable_candidate":
            visibility = "metadata_only"
            basis = "受控原件路径不可读，自动降级为metadata_only"
        decision = {
            "allowlist_entry_id": None,
            "source_id": source_id,
            "asset_id": (asset or {}).get("asset_id"),
            "document_id": None,
            "visibility": visibility,
            "decision_basis": basis,
            "decision_source_file": source.get("source_record_file") or "20_补充政策文献与国家因子库接收台账.xlsx:01_政策标准PDF",
            "decision_source_row": None,
            "mapping_method": mapping_method,
            "decided_at": DECISION_DATE,
            "allowlist_version": ALLOWLIST_VERSION,
            "expected_sha256": str((asset or {}).get("_expected_sha256") or "") or None,
            "actual_sha256": (asset or {}).get("_actual_sha256"),
            "hash_status": hash_status,
            "date_uncertain": _record_date_uncertain(source.get("effective_at"), source.get("published_at"), source.get("version")),
            "source": source,
            "asset": asset,
        }
        decision["allowlist_entry_id"] = self._admission_decision_id(decision)
        return decision

    @staticmethod
    def _admission_decision_id(decision: dict[str, Any]) -> str:
        """Return an immutable ID for one admission decision snapshot.

        Visibility and hash state are part of the identity. If the same source
        later moves from searchable to blocked, the new decision must not be
        swallowed by INSERT OR IGNORE under the old source-only ID.
        """
        source = decision.get("source") or {}
        identity = {
            "source_id": decision.get("source_id"),
            "source_record_hash": _sha256(_json(source)),
            "allowlist_version": decision.get("allowlist_version"),
            "visibility": decision.get("visibility"),
            "decision_basis": decision.get("decision_basis"),
            "asset_id": decision.get("asset_id"),
            "document_id": decision.get("document_id"),
            "expected_sha256": decision.get("expected_sha256"),
            "actual_sha256": decision.get("actual_sha256"),
            "hash_status": decision.get("hash_status"),
            "mapping_method": decision.get("mapping_method"),
        }
        return f"allow-{_slug(str(decision.get('source_id') or 'source'))}-{_sha256(_json(identity))[:16]}"

    def _manifest(self) -> dict[str, Any]:
        source_rows = self._source_rows()
        asset_rows = self._asset_rows()
        assets_by_source: dict[str, list[dict[str, Any]]] = defaultdict(list)
        assets_by_hash: dict[str, dict[str, Any]] = {}
        for row in asset_rows:
            source_id = self._asset_source_id(row)
            assets_by_source[source_id].append(row)
            file_hash = str(row.get("SHA-256") or "").lower()
            if file_hash:
                assets_by_hash[file_hash] = row
            if source_id not in source_rows:
                source_rows[source_id] = {
                    "source_id": source_id,
                    "title": row.get("文件名"),
                    "publisher": "待补来源映射",
                    "version": None,
                    "official_url": None,
                    "verification_status": "部分核验",
                    "source_identity_status": "待人工映射",
                    "fulltext_status": row.get("文本状态"),
                    "admission_status": "候选；不得自动升级为正式规则",
                    "role": "other",
                    "use_boundary": row.get("执行边界") or "仅作候选知识资产",
                    "source_record_file": "20_补充政策文献与国家因子库接收台账.xlsx:01_政策标准PDF",
                }

        decisions: list[dict[str, Any]] = []
        source_entries: list[dict[str, Any]] = []
        documents: list[dict[str, Any]] = []
        chunks: list[dict[str, Any]] = []
        reference_leaks: list[dict[str, Any]] = []
        parse_errors: list[dict[str, Any]] = []
        duplicate_hashes = {key: len(value) for key, value in Counter(str(row.get("SHA-256") or "") for row in asset_rows if row.get("SHA-256")).items() if value > 1}
        hash_issues = [
            {
                "asset_id": row.get("asset_id"),
                "expected_sha256": row.get("_expected_sha256"),
                "actual_sha256": row.get("_actual_sha256"),
                "status": row.get("_hash_status"),
            }
            for row in asset_rows
            if row.get("_hash_status") != "match"
        ]

        # The only automatic source-to-file fallback is the exact hash binding
        # frozen by the 14 copper clause records.
        clause_rows = _load_rows(self.assets.intake_ledger, "04_铜行业条款核验")
        clause_hash_to_asset = {
            str(row.get("文件SHA-256") or "").lower(): assets_by_hash.get(str(row.get("文件SHA-256") or "").lower())
            for row in clause_rows
            if row.get("文件SHA-256")
        }
        for source_id in sorted(source_rows):
            source = source_rows[source_id]
            candidates = assets_by_source.get(source_id, [])
            asset = candidates[0] if candidates else None
            mapping_method = "source_id_exact" if asset else "unmapped"
            if source_id in {"STD-003", "STD-004"} and clause_hash_to_asset:
                hashes = [str(row.get("文件SHA-256") or "").lower() for row in clause_rows if row.get("source_id") == source_id]
                exact_assets = [clause_hash_to_asset.get(item) for item in hashes if clause_hash_to_asset.get(item)]
                if exact_assets:
                    asset = exact_assets[0]
                    mapping_method = "source_id_plus_document_sha256"
            decision = self._decision_for_source(source, asset, mapping_method=mapping_method)
            decision["decision_source_row"] = next((index for index, row in enumerate(self._source_rows().values(), start=2) if row.get("source_id") == source_id), None)
            if asset and decision["visibility"] == "searchable_candidate":
                document_hash = str(asset.get("SHA-256") or "").lower()
                document_id = f"doc-{_slug(source_id)}-{document_hash[:16]}"
                decision["document_id"] = document_id
                pages = _read_pdf_pages(asset["_path"])
                if not pages:
                    parse_errors.append({"source_id": source_id, "asset_id": asset.get("asset_id"), "reason": "无法读取正文"})
                    decision["visibility"] = "metadata_only"
                else:
                    documents.append(self._document(source, asset, document_id, pages, decision["visibility"]))
                    chunks.extend(self._page_chunks(source, asset, document_id, pages))
            elif asset:
                document_hash = str(asset.get("SHA-256") or "").lower()
                if document_hash:
                    document_id = f"doc-{_slug(source_id)}-{document_hash[:16]}"
                    decision["document_id"] = document_id
                    documents.append(self._document(source, asset, document_id, [], decision["visibility"]))
            # The document ID and final parse visibility are also part of the
            # decision snapshot. Recompute after those fields are known.
            decision["allowlist_entry_id"] = self._admission_decision_id(decision)
            public_source = self._source_entry(decision, source)
            if has_reference_marker(public_source):
                reference_leaks.append({"kind": "source", "source_id": source_id})
            source_entries.append(public_source)
            decisions.append({key: value for key, value in decision.items() if key not in {"source", "asset"}})

        # Copper clauses have a stronger exact source_id + SHA-256 binding than
        # the generic asset rows and therefore receive precise candidate chunks.
        for row in clause_rows:
            source_id = str(row.get("source_id") or "")
            asset = clause_hash_to_asset.get(str(row.get("文件SHA-256") or "").lower())
            source = source_rows.get(source_id)
            if not source or not asset or source_id not in SEARCHABLE_SOURCE_IDS:
                continue
            document_id = f"doc-{_slug(source_id)}-{str(asset.get('SHA-256') or '').lower()[:16]}"
            excerpt = _short_text(row.get("逐字摘录"), 2400)
            if not excerpt or has_reference_marker(excerpt):
                if has_reference_marker(excerpt):
                    reference_leaks.append({"kind": "copper_clause", "source_id": source_id, "clause_id": row.get("clause_evidence_id")})
                continue
            locator = f"{row.get('章节/条款') or ''}；表{row.get('表号') or ''}；PDF第{row.get('PDF页码') or ''}页"
            chunks.append(
                self._chunk(
                    source,
                    asset,
                    document_id,
                    chunk_type="clause",
                    title=row.get("主题") or source.get("title"),
                    section_title=row.get("章节/条款"),
                    locator=locator,
                    page_start=int(row.get("PDF页码") or 0) or None,
                    text=excerpt,
                    verification_status="candidate_evidence_verified",
                    use_boundary=row.get("使用边界") or source.get("use_boundary") or "候选证据；不得自动评分",
                )
            )

        # Twelve Chinese papers are admitted only as governed research
        # fragments. English papers remain blocked because no full text exists.
        for row in self._literature_rows("01_中文文献"):
            literature_id = str(row.get("literature_id") or "")
            path = _literature_pdf_path(self.project_root, row.get("full_text_availability"))
            hash_match = re.search(r"SHA-256=([0-9a-fA-F]{64})", str(row.get("full_text_availability") or ""))
            file_hash = hash_match.group(1).lower() if hash_match else ""
            actual_hash = _sha256(path.read_bytes()) if path and path.is_file() else None
            literature_hash_status = "missing" if not file_hash else "unreadable" if actual_hash is None else "match" if actual_hash == file_hash else "mismatch"
            if literature_hash_status != "match":
                hash_issues.append({"asset_id": literature_id, "expected_sha256": file_hash or None, "actual_sha256": actual_hash, "status": literature_hash_status})
            if literature_hash_status != "match":
                path = None
            research_scope = list(RESEARCH_INDUSTRY_SCOPE.get(literature_id, []))
            source = {
                "source_id": literature_id,
                "title": row.get("full_title"),
                "publisher": row.get("journal_publisher"),
                "document_no_or_standard_no": row.get("doi"),
                "version": row.get("year"),
                "published_at": row.get("year"),
                "effective_at": None,
                "official_url": row.get("official_url"),
                "verification_status": row.get("evidence_level") or row.get("verification_status") or "研究证据初核",
                "source_identity_status": "已核验",
                "source_role": "research_literature",
                "role": "research_literature",
                "industry_scope": research_scope,
                "use_boundary": row.get("unsupported_conclusions") or "研究证据；不替代政策、标准、因子或评分规则",
            }
            research_mapping_method = "literature_id_exact+governance_global_scope"
            research_asset = {"_path": path, "_relative_path": path.relative_to(self.project_root).as_posix() if path else None, "SHA-256": file_hash or None, "_expected_sha256": file_hash or None, "_actual_sha256": actual_hash, "_hash_status": literature_hash_status, "文件名": path.name if path else None, "asset_id": None}
            research_decision = {
                "allowlist_entry_id": None,
                "source_id": literature_id,
                "asset_id": None,
                "document_id": f"doc-{_slug(literature_id)}-{file_hash[:16]}",
                "visibility": "searchable_candidate" if path and path.is_file() and literature_hash_status == "match" else "blocked",
                "decision_basis": RESEARCH_SCOPE_BASIS + ("；全文SHA-256复核一致，仅导入治理确认的摘要/结论片段" if path else f"；全文正文SHA-256{literature_hash_status}，正文阻断"),
                "mapping_method": research_mapping_method,
                "source": source,
                "asset": research_asset,
                "expected_sha256": file_hash or None,
                "actual_sha256": actual_hash,
                "hash_status": literature_hash_status,
                "date_uncertain": _record_date_uncertain(source.get("published_at"), source.get("version")),
                "allowlist_version": ALLOWLIST_VERSION,
            }
            research_decision["allowlist_entry_id"] = self._admission_decision_id(research_decision)
            source_uid = f"src-{_slug(literature_id)}-{_sha256(_json(source))[:16]}"
            source_entry = self._source_entry(
                research_decision,
                source,
            )
            source_entry["source_uid"] = source_uid
            source_entries.append(source_entry)
            research_decision.update({
                "document_id": source_entry["document_id"],
                "visibility": source_entry["visibility"],
                "decision_source_file": "14_中文与国际学术文献证据表.xlsx:01_中文文献",
                "decision_source_row": None,
                "decided_at": DECISION_DATE,
            })
            research_decision["allowlist_entry_id"] = self._admission_decision_id(research_decision)
            decisions.append({key: value for key, value in research_decision.items() if key not in {"source", "asset", "expected_sha256", "actual_sha256", "hash_status", "date_uncertain"}})
            if path and path.is_file() and literature_hash_status == "match":
                pages = _read_pdf_pages(path)
                document = self._document(source, research_asset, source_entry["document_id"], pages, "searchable_candidate")
                document["source_uid"] = source_entry["source_uid"]
                documents.append(document)
                for page_number, label, text in _extract_research_pages(pages):
                    if has_reference_marker(text):
                        reference_leaks.append({"kind": "literature", "source_id": literature_id, "page": page_number})
                        continue
                    chunks.append(
                        self._chunk(
                            source,
                            research_asset,
                            source_entry["document_id"],
                            chunk_type="research_fragment",
                            title=row.get("full_title"),
                            section_title=label,
                            locator=f"PDF第{page_number}页；{label}",
                            page_start=page_number,
                            text=_short_text(text, 2400),
                            verification_status="research_evidence_verified",
                            use_boundary=source["use_boundary"],
                        )
                    )

        candidates = _load_rows(self.assets.candidates, "01_候选语料")
        governance: list[dict[str, Any]] = []
        governance_seen: set[str] = set()
        for row in candidates:
            item_id = str(row.get("item_id") or "")
            source_type = str(row.get("source_type") or "")
            source_record_id = str(row.get("source_record_id") or item_id).strip()
            if not source_record_id or source_record_id in governance_seen:
                continue
            governance_seen.add(source_record_id)
            status = "blocked" if source_type == "factor_candidate" else "diagnostic_only"
            text = _short_text(row.get("retrieval_text") or row.get("description"), 2400)
            if has_reference_marker(row):
                reference_leaks.append({"kind": "governance", "record_id": item_id})
            governance.append(
                {
                    # 12号清单的 source_record_id is the stable governance
                    # identity. Do not manufacture a second record ID from a
                    # row hash, otherwise the 11 GOV-IND rows are duplicated.
                    "governance_record_id": source_record_id,
                    "record_type": source_type or "candidate",
                    "source_record_id": source_record_id,
                    "industry": normalize_industry(row.get("industry")),
                    "status": status,
                    "issue_summary": row.get("exclusion_reason") or "12号治理记录，不进入普通索引",
                    "required_action": "补齐正式来源、版本、哈希、适用范围和精确定位后重新审查",
                    "source_workbook": "12_M5候选语料与切片清单.xlsx:01_候选语料",
                    "source_row": None,
                    "content_hash": row.get("content_hash") or _sha256(text),
                    "text": text,
                    "chunk_id": row.get("chunk_id"),
                    "item_id": item_id,
                    "visibility": "diagnostic_only",
                    "updated_at": DECISION_DATE,
                }
            )

        # Seven English entries are registered as blocked metadata, never as
        # searchable research fragments.
        english = self._literature_rows("02_English literature")
        for row in english:
            literature_id = str(row.get("literature_id") or "")
            source = {
                "source_id": literature_id,
                "title": row.get("full_title"),
                "publisher": row.get("journal_publisher"),
                "document_no_or_standard_no": row.get("doi"),
                "version": row.get("year"),
                "official_url": row.get("official_url"),
                "verification_status": row.get("verification_status") or "仅元数据核验",
                "source_role": "research_literature",
                "role": "research_literature",
                "use_boundary": row.get("unsupported_conclusions") or "英文全文待补；不得作为正文证据",
            }
            source_entries.append(self._source_entry({
                "source_id": literature_id,
                "asset_id": None,
                "document_id": None,
                "visibility": "blocked",
                "allowlist_entry_id": None,
                "decision_basis": "英文全文尚未取得，正文阻断",
                "mapping_method": "literature_id_exact",
                "source": source,
                "asset": None,
            }, source))
            english_decision = {
                "allowlist_entry_id": None,
                "source_id": literature_id,
                "asset_id": None,
                "document_id": None,
                "visibility": "blocked",
                "decision_basis": "英文全文尚未取得，正文阻断",
                "decision_source_file": "14_中文与国际学术文献证据表.xlsx:02_English literature",
                "decision_source_row": None,
                "mapping_method": "literature_id_exact",
                "decided_at": DECISION_DATE,
                "allowlist_version": ALLOWLIST_VERSION,
                "source": source,
                "asset": None,
                "expected_sha256": None,
                "actual_sha256": None,
                "hash_status": "missing",
            }
            english_decision["allowlist_entry_id"] = self._admission_decision_id(english_decision)
            source_entries[-1]["allowlist_entry_id"] = english_decision["allowlist_entry_id"]
            decisions.append({key: value for key, value in english_decision.items() if key not in {"source", "asset", "expected_sha256", "actual_sha256", "hash_status"}})

        if has_reference_marker(chunks) or has_reference_marker(documents) or has_reference_marker(source_entries):
            reference_leaks.append({"kind": "manifest_recursive_scan"})

        # A content hash of stable, path-free manifest fields makes identical
        # builds reuse one index version while preserving old versions if inputs change.
        stable_manifest = {
            "allowlist_version": ALLOWLIST_VERSION,
            "ranking_config_version": RANKING_CONFIG_VERSION,
            "tokenizer_version": TOKENIZER_VERSION,
            "industry_mapping_version": INDUSTRY_MAPPING_VERSION,
            "sources": [{key: value for key, value in item.items() if key != "asset"} for item in source_entries],
            "documents": [{key: value for key, value in item.items() if key != "_path"} for item in documents],
            "chunks": chunks,
            "governance": [{key: value for key, value in item.items() if key != "text"} for item in governance],
            "hash_issues": hash_issues,
        }
        manifest_hash = _sha256(_json(stable_manifest))
        visibility_counts = Counter(item["visibility"] for item in source_entries)
        governance_counts = Counter(item["status"] for item in governance)
        low_text_count = sum(1 for row in asset_rows if "扫描/低文本" in str(row.get("文本状态") or ""))
        invalid_count = sum(1 for row in asset_rows if "0页" in str(row.get("接收结论") or "") or "不可解析" in str(row.get("文本状态") or ""))
        stats = {
            "discovered_source_records": len(source_rows),
            "source_entries": len(source_entries),
            "searchable_candidate": visibility_counts.get("searchable_candidate", 0),
            "metadata_only": visibility_counts.get("metadata_only", 0),
            "diagnostic_only": governance_counts.get("diagnostic_only", 0),
            "blocked": visibility_counts.get("blocked", 0) + governance_counts.get("blocked", 0),
            "governance_records": len(governance),
            "documents": len({item["document_id"] for item in documents}),
            "chunks": len({item["chunk_id"] for item in chunks}),
            "copper_clause_chunks": sum(1 for item in chunks if item["chunk_type"] == "clause"),
            "research_fragment_chunks": sum(1 for item in chunks if item["chunk_type"] == "research_fragment"),
            "duplicate_hash_groups": len(duplicate_hashes),
            "hash_issue_count": len(hash_issues),
            "low_text_pdf_count": low_text_count,
            "invalid_pdf_count": invalid_count,
            "english_fulltext_blocked": len(english),
            "factor_candidate_blocked": sum(1 for item in governance if item["record_type"] == "factor_candidate"),
            "industry_count": len(INDUSTRIES),
            "reference_conclusion_hits": len(reference_leaks),
        }
        errors = list(parse_errors)
        errors.extend({"kind": "reference_leak", **item} for item in reference_leaks)
        return {
            "allowlist_version": ALLOWLIST_VERSION,
            "manifest_hash": manifest_hash,
            "index_version_id": f"m5-index-{manifest_hash[:16]}",
            "ranking_config_version": RANKING_CONFIG_VERSION,
            "ranking_config_hash": _sha256(_json({
                "version": RANKING_CONFIG_VERSION,
                "role_order": sorted((key, value) for key, value in {"official_standard": 0, "official_policy": 1, "regulatory_guidance": 2, "official_methodology": 3, "research_literature": 4, "other": 5}.items()),
                "bm25_weights": list(BM25_WEIGHTS),
            })),
            "tokenizer_version": TOKENIZER_VERSION,
            "industry_mapping_version": INDUSTRY_MAPPING_VERSION,
            "source_entries": source_entries,
            "decisions": decisions,
            "documents": documents,
            "chunks": chunks,
            "governance": governance,
            "duplicates": duplicate_hashes,
            "parse_errors": parse_errors,
            "reference_leaks": reference_leaks,
            "hash_issues": hash_issues,
            "blocking_errors": errors,
            "stats": stats,
        }

    @staticmethod
    def _source_entry(decision: dict[str, Any], source: dict[str, Any]) -> dict[str, Any]:
        source_id = str(source.get("source_id") or decision.get("source_id") or "")
        effective_at = source.get("effective_at") or source.get("effective_date")
        published_at = source.get("published_at") or source.get("year")
        version = source.get("version") or source.get("year")
        source_uid = f"src-{_slug(source_id)}-{_sha256(_json({key: value for key, value in source.items() if not str(key).startswith('_')}))[:16]}"
        return {
            "source_uid": source_uid,
            "source_id": source_id,
            "canonical_source_id": source.get("canonical_source_id") or source_id,
            "title": source.get("title") or source.get("full_title") or source_id,
            "publisher": source.get("publisher"),
            "document_no_or_standard_no": source.get("document_no_or_standard_no") or source.get("doi"),
            "source_role": source.get("source_role") or source.get("role") or "other",
            "version": version,
            "published_at": published_at,
            "effective_at": effective_at,
            "expires_at": source.get("expires_at"),
            "region": source.get("region") or "全国",
            "industry_scope": list(source.get("industry_scope")) if isinstance(source.get("industry_scope"), list) else industry_scope_for_source(source_id, source.get("role")),
            "official_url": public_url(source.get("official_url")),
            "verification_status": source.get("verification_status") or "部分核验",
            "admission_status": source.get("admission_status") or "候选；不得自动升级",
            "visibility": decision.get("visibility", "blocked"),
            "use_boundary": source.get("use_boundary") or "候选证据；不得自动执行规则、因子、评分或授信",
            "supersedes_source_id": source.get("supersedes_source_id"),
            "asset_id": decision.get("asset_id"),
            "document_id": decision.get("document_id"),
            "mapping_method": decision.get("mapping_method"),
            "record_hash": _sha256(_json(source)),
            "created_at": DECISION_DATE,
            "updated_at": DECISION_DATE,
            "date_uncertain": bool(decision.get("date_uncertain")) or _record_date_uncertain(effective_at, published_at, version),
            "asset": decision.get("asset"),
        }

    @staticmethod
    def _document(source: dict[str, Any], asset: dict[str, Any], document_id: str, pages: list[str], visibility: str) -> dict[str, Any]:
        return {
            "document_id": document_id,
            "source_id": source.get("source_id"),
            "source_uid": f"src-{_slug(str(source.get('source_id') or ''))}-{_sha256(_json(source))[:16]}",
            "file_name": asset.get("文件名"),
            "relative_path": asset.get("_relative_path") or f"asset:{asset.get('asset_id')}",
            "sha256": str(asset.get("SHA-256") or "").lower(),
            "mime_type": "application/pdf",
            "page_count": len(pages) if pages else asset.get("页数"),
            "text_status": asset.get("文本状态"),
            "parser_version": "pdf-text-v1",
            "received_at": DECISION_DATE,
            "is_original_immutable": True,
            "visibility": visibility,
        }

    @staticmethod
    def _chunk(
        source: dict[str, Any],
        asset: dict[str, Any],
        document_id: str,
        *,
        chunk_type: str,
        title: Any,
        section_title: Any,
        locator: str,
        page_start: int | None,
        text: str,
        verification_status: str,
        use_boundary: str,
    ) -> dict[str, Any]:
        content_hash = _sha256(text)
        chunk_id = f"chunk-{_sha256(_json([source.get('source_id'), str(asset.get('SHA-256') or '').lower(), locator, content_hash]))[:24]}"
        industry_scope = list(source.get("industry_scope")) if isinstance(source.get("industry_scope"), list) else industry_scope_for_source(str(source.get("source_id") or ""), source.get("role"))
        return {
            "chunk_id": chunk_id,
            "document_id": document_id,
            "source_id": source.get("source_id"),
            "chunk_type": chunk_type,
            "title": title or source.get("title"),
            "section_title": section_title,
            "page_start": page_start,
            "page_end": page_start,
            "clause_no": None,
            "table_no": None,
            "locator": locator,
            "text": text,
            "normalized_search_text": normalize_search_text(source.get("source_id"), source.get("title"), source.get("document_no_or_standard_no"), title, section_title, text, *industry_scope),
            "content_hash": content_hash,
            "verification_status": verification_status,
            "visibility": "searchable_candidate",
            "use_boundary": use_boundary,
            "industry_scope": industry_scope,
            "source_role": source.get("source_role") or source.get("role") or "other",
            "official_url": public_url(source.get("official_url")),
            "publisher": source.get("publisher"),
            "version": source.get("version") or source.get("year"),
        }

    def _page_chunks(self, source: dict[str, Any], asset: dict[str, Any], document_id: str, pages: list[str]) -> list[dict[str, Any]]:
        output: list[dict[str, Any]] = []
        for page_number, text in enumerate(pages, start=1):
            for segment_number, segment in enumerate(_split_page(text), start=1):
                output.append(
                    self._chunk(
                        source,
                        asset,
                        document_id,
                        chunk_type="page",
                        title=source.get("title"),
                        section_title=None,
                        locator=f"PDF第{page_number}页" + (f"；片段{segment_number}" if len(_split_page(text)) > 1 else ""),
                        page_start=page_number,
                        text=segment,
                        verification_status=source.get("verification_status") or "candidate_evidence",
                        use_boundary=source.get("use_boundary") or "候选证据；不得自动执行规则、因子、评分或授信",
                    )
                )
        return output

    def dry_run(self) -> dict[str, Any]:
        manifest = self._manifest()
        return {
            "status": "blocked" if manifest["blocking_errors"] else "ready",
            "allowlist_version": manifest["allowlist_version"],
            "manifest_hash": manifest["manifest_hash"],
            "index_version_id": manifest["index_version_id"],
            "stats": manifest["stats"],
            "visibility_counts": dict(Counter(item["visibility"] for item in manifest["source_entries"])),
            "governance_status_counts": dict(Counter(item["status"] for item in manifest["governance"])),
            "duplicate_hashes": manifest["duplicates"],
            "hash_issues": manifest["hash_issues"],
            "unparseable_files": manifest["parse_errors"],
            "reference_conclusion_hits": manifest["reference_leaks"],
            "blocking_errors": manifest["blocking_errors"],
            "write_plan": {
                "sources": len(manifest["source_entries"]),
                "documents": len(manifest["documents"]),
                "chunks": len(manifest["chunks"]),
                "governance_records": len(manifest["governance"]),
                "ordinary_index_records": sum(1 for item in manifest["chunks"] if item["visibility"] == "searchable_candidate") + sum(1 for item in manifest["source_entries"] if item["visibility"] in {"searchable_candidate", "metadata_only"}),
            },
            "notice": "dry-run只读取治理台账和受控原件，不写入SQLite索引；原始数据为命题方脱敏模拟数据。",
        }

    def rebuild(self) -> dict[str, Any]:
        manifest = self._manifest()
        if manifest["blocking_errors"]:
            raise KnowledgeBuildBlocked("M5索引构建被dry-run硬门禁阻断", self.dry_run_from_manifest(manifest))
        index_id = manifest["index_version_id"]
        with self.domain_store._connection() as connection:
            existing = connection.execute("SELECT * FROM knowledge_index_versions WHERE index_version_id = ?", (index_id,)).fetchone()
            if existing is not None and existing["build_status"] == "built" and existing["manifest_hash"] == manifest["manifest_hash"]:
                # Reusing a frozen manifest is deliberately a no-op. This
                # preserves member rows, retrieval replay and admission
                # history instead of rewriting an index in place.
                connection.execute("UPDATE knowledge_index_versions SET is_current = 0 WHERE is_current = 1 AND index_version_id != ?", (index_id,))
                connection.execute("UPDATE knowledge_index_versions SET is_current = 1 WHERE index_version_id = ?", (index_id,))
                return self.index_status()
            connection.execute("UPDATE knowledge_index_versions SET is_current = 0 WHERE is_current = 1")
            connection.execute(
                """INSERT OR IGNORE INTO knowledge_index_versions(
                    index_version_id,allowlist_version,manifest_hash,extractor_version,chunker_version,
                    tokenizer_version,built_at,source_count,document_count,chunk_count,
                    searchable_candidate_count,metadata_only_count,diagnostic_only_count,blocked_count,
                    ranking_config_hash,build_status,error_summary,is_current
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)""",
                (
                    index_id,
                    manifest["allowlist_version"],
                    manifest["manifest_hash"],
                    "pdf-text-v1",
                    "page-1200-v1",
                    manifest["tokenizer_version"],
                    _now(),
                    len(manifest["source_entries"]),
                    len({item["document_id"] for item in manifest["documents"]}),
                    len({item["chunk_id"] for item in manifest["chunks"]}),
                    manifest["stats"]["searchable_candidate"],
                    manifest["stats"]["metadata_only"],
                    manifest["stats"]["diagnostic_only"],
                    manifest["stats"]["blocked"],
                    manifest["ranking_config_hash"],
                    "built",
                    "",
                ),
            )
            connection.execute("DELETE FROM knowledge_index_members WHERE index_version_id = ?", (index_id,))
            if self.fts_available:
                connection.execute("DELETE FROM knowledge_fts WHERE index_version_id = ?", (index_id,))
            source_uid_by_id = {source["source_id"]: source["source_uid"] for source in manifest["source_entries"]}
            for source in manifest["source_entries"]:
                source_uid = source["source_uid"]
                connection.execute(
                    """INSERT OR IGNORE INTO knowledge_sources(
                        source_uid,source_id,canonical_source_id,title,publisher,document_no_or_standard_no,
                        source_role,version,published_at,effective_at,expires_at,region,industry_scope_json,
                        official_url,verification_status,admission_status,visibility,use_boundary,
                        supersedes_source_id,asset_id,document_id,mapping_method,record_hash,created_at,updated_at,date_uncertain
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        source_uid, source["source_id"], source["canonical_source_id"], source["title"], source["publisher"],
                        source["document_no_or_standard_no"], source["source_role"], source["version"], source["published_at"],
                        source["effective_at"], source["expires_at"], source["region"], _json(source["industry_scope"]), source["official_url"],
                        source["verification_status"], source["admission_status"], source["visibility"], source["use_boundary"],
                        source["supersedes_source_id"], source["asset_id"], source["document_id"], source["mapping_method"],
                        source["record_hash"], source["created_at"], source["updated_at"],
                        int(bool(source.get("date_uncertain"))),
                    ),
                )
                if source["visibility"] in {"searchable_candidate", "metadata_only"}:
                    entry_id = f"source_metadata:{source_uid}"
                    connection.execute(
                        """INSERT OR REPLACE INTO knowledge_index_members(
                            index_version_id,entry_id,entry_type,source_uid,source_id,chunk_id,visibility,
                            verification_status,use_boundary,ranking_config_hash,included_at
                        ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                        (index_id, entry_id, "source_metadata", source_uid, source["source_id"], None, source["visibility"], source["verification_status"], source["use_boundary"], manifest["ranking_config_hash"], _now()),
                    )
                    if self.fts_available:
                        connection.execute("INSERT INTO knowledge_fts(entry_id,index_version_id,source_id,title,standard_no,section_title,body,normalized_search_text) VALUES(?,?,?,?,?,?,?,?)", (entry_id, index_id, source["source_id"], source["title"], source["document_no_or_standard_no"] or "", "", "", normalize_search_text(source["source_id"], source["title"], source["document_no_or_standard_no"])))
            for document in manifest["documents"]:
                connection.execute(
                    """INSERT OR IGNORE INTO knowledge_documents(
                        document_id,source_id,source_uid,file_name,relative_path,sha256,mime_type,page_count,
                        text_status,parser_version,received_at,is_original_immutable,visibility
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (document["document_id"], document["source_id"], document["source_uid"], document["file_name"], document["relative_path"], document["sha256"], document["mime_type"], document["page_count"], document["text_status"], document["parser_version"], document["received_at"], int(document["is_original_immutable"]), document["visibility"]),
                )
            for chunk in manifest["chunks"]:
                connection.execute(
                    """INSERT OR IGNORE INTO knowledge_chunks(
                        chunk_id,document_id,source_id,chunk_type,title,section_title,page_start,page_end,
                        clause_no,table_no,locator,text,normalized_search_text,content_hash,verification_status,
                        visibility,use_boundary,industry_scope_json,source_role,official_url,publisher,version
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (chunk["chunk_id"], chunk["document_id"], chunk["source_id"], chunk["chunk_type"], chunk["title"], chunk["section_title"], chunk["page_start"], chunk["page_end"], chunk["clause_no"], chunk["table_no"], chunk["locator"], chunk["text"], chunk["normalized_search_text"], chunk["content_hash"], chunk["verification_status"], chunk["visibility"], chunk["use_boundary"], _json(chunk["industry_scope"]), chunk["source_role"], chunk["official_url"], chunk["publisher"], chunk["version"]),
                )
                entry_id = f"chunk:{chunk['chunk_id']}"
                connection.execute(
                    """INSERT OR REPLACE INTO knowledge_index_members(
                        index_version_id,entry_id,entry_type,source_uid,source_id,chunk_id,visibility,
                        verification_status,use_boundary,ranking_config_hash,included_at
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                    (index_id, entry_id, "chunk", source_uid_by_id.get(chunk["source_id"], f"src-{_slug(str(chunk['source_id']))}-{_sha256(str(chunk['source_id']))[:16]}"), chunk["source_id"], chunk["chunk_id"], chunk["visibility"], chunk["verification_status"], chunk["use_boundary"], manifest["ranking_config_hash"], _now()),
                )
                if self.fts_available:
                    connection.execute("INSERT INTO knowledge_fts(entry_id,index_version_id,source_id,title,standard_no,section_title,body,normalized_search_text) VALUES(?,?,?,?,?,?,?,?)", (entry_id, index_id, chunk["source_id"], chunk["title"], "", chunk["section_title"] or "", chunk["text"], chunk["normalized_search_text"]))
            for decision in manifest["decisions"]:
                previous_decision = connection.execute(
                    "SELECT decision_id FROM knowledge_admission_decisions WHERE source_id = ? ORDER BY decided_at DESC, decision_id DESC LIMIT 1",
                    (decision["source_id"],),
                ).fetchone()
                supersedes = previous_decision["decision_id"] if previous_decision and previous_decision["decision_id"] != decision["allowlist_entry_id"] else None
                connection.execute(
                    """INSERT OR IGNORE INTO knowledge_admission_decisions(
                        decision_id,allowlist_version,source_id,asset_id,document_id,visibility,decision_basis,
                        decision_source_file,decision_source_row,mapping_method,decided_at,supersedes_decision_id
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (decision["allowlist_entry_id"], decision["allowlist_version"], decision["source_id"], decision["asset_id"], decision["document_id"], decision["visibility"], decision["decision_basis"], decision["decision_source_file"], decision["decision_source_row"], decision["mapping_method"], decision["decided_at"], supersedes),
                )
            for record in manifest["governance"]:
                # Older local M5 builds used a row-hash-prefixed governance ID
                # and therefore duplicated the 11 GOV-IND records. The
                # governance table is a current registry (not an index
                # history table), so remove only non-canonical duplicates by
                # stable source_record_id before inserting the canonical row.
                connection.execute(
                    "DELETE FROM knowledge_governance_records WHERE source_record_id = ? AND governance_record_id != ?",
                    (record["source_record_id"], record["governance_record_id"]),
                )
                connection.execute(
                    """INSERT OR IGNORE INTO knowledge_governance_records(
                        governance_record_id,record_type,source_record_id,industry,status,issue_summary,
                        required_action,source_workbook,source_row,content_hash,text,chunk_id,visibility,updated_at
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (record["governance_record_id"], record["record_type"], record["source_record_id"], record["industry"], record["status"], record["issue_summary"], record["required_action"], record["source_workbook"], record["source_row"], record["content_hash"], record["text"], record["chunk_id"], record["visibility"], record["updated_at"]),
                )
            connection.execute("UPDATE knowledge_index_versions SET is_current = 1 WHERE index_version_id = ?", (index_id,))
        return self.index_status()

    @staticmethod
    def dry_run_from_manifest(manifest: dict[str, Any]) -> dict[str, Any]:
        return {
            "status": "blocked" if manifest["blocking_errors"] else "ready",
            "allowlist_version": manifest["allowlist_version"],
            "manifest_hash": manifest["manifest_hash"],
            "index_version_id": manifest["index_version_id"],
            "stats": manifest["stats"],
            "visibility_counts": dict(Counter(item["visibility"] for item in manifest["source_entries"])),
            "governance_status_counts": dict(Counter(item["status"] for item in manifest["governance"])),
            "duplicate_hashes": manifest["duplicates"],
            "hash_issues": manifest["hash_issues"],
            "unparseable_files": manifest["parse_errors"],
            "reference_conclusion_hits": manifest["reference_leaks"],
            "blocking_errors": manifest["blocking_errors"],
        }

    def index_status(self) -> dict[str, Any]:
        with self.domain_store._connection() as connection:
            row = connection.execute("SELECT * FROM knowledge_index_versions WHERE is_current = 1 ORDER BY built_at DESC LIMIT 1").fetchone()
            if row is None:
                return {"available": False, "fts5_available": self.fts_available, "allowlist_version": ALLOWLIST_VERSION}
            return {
                "available": True,
                "fts5_available": self.fts_available,
                "index_version_id": row["index_version_id"],
                "allowlist_version": row["allowlist_version"],
                "manifest_hash": row["manifest_hash"],
                "built_at": row["built_at"],
                "source_count": row["source_count"],
                "document_count": row["document_count"],
                "chunk_count": row["chunk_count"],
                "searchable_candidate_count": row["searchable_candidate_count"],
                "metadata_only_count": row["metadata_only_count"],
                "diagnostic_only_count": row["diagnostic_only_count"],
                "blocked_count": row["blocked_count"],
                "ranking_config_hash": row["ranking_config_hash"],
                "build_status": row["build_status"],
                "notice": "索引只包含记录级准入后的本地知识候选；不包含转型规划结论、企业附件或具体排放因子值。",
            }

    def ensure_index(self) -> dict[str, Any]:
        status = self.index_status()
        if status.get("available"):
            return status
        raise KnowledgeIndexNotReady("knowledge_index_not_ready: 本地知识索引尚未构建，请由管理员执行受控构建")

    def _run_industry(self, assessment_run_id: str) -> tuple[dict[str, Any], str | None]:
        run = self.domain_store.get_assessment_run(assessment_run_id)
        if self.analysis_loader is None:
            return run, None
        _loaded_run, analysis = self.analysis_loader(assessment_run_id)
        basic = (analysis.get("input_data") or {}).get("basic_info") or {}
        return run, normalize_industry(basic.get("行业"))

    @staticmethod
    def _redacted_query(query: str) -> str:
        return re.sub(r"TF\d{4,}|\d+(?:\.\d+)?", "[已脱敏]", query[:180])

    @staticmethod
    def _query_guard(query: str) -> str | None:
        """Apply deterministic negative-query gates before ordinary retrieval."""
        text = str(query or "").casefold()
        if has_reference_marker(query) or "参考结论" in text or "参考对照" in text:
            return "参考对照层字段不允许作为知识检索输入"
        if ("铜企业" in text and "钢铁" in text) or "跨行业" in text or "专属规范" in text:
            return "行业知识必须按当前企业行业范围过滤，禁止跨行业套用专属规范"
        if any(marker in text for marker in ("直接执行候选规则", "候选规则阈值", "执行阈值", "自动执行规则")):
            return "候选规则未完成正式来源、版本和定位核验，禁止自动执行"
        if any(marker in text for marker in ("未经核准排放因子", "直接调用排放因子", "调用具体排放因子", "排放因子值")):
            return "M5只登记因子来源和适用性，不自动调用具体排放因子值"
        if "无正式url" in text or "只有标题" in text or "标题无" in text:
            return "来源只有标题或缺少正式URL，正式证据返回被阻断"
        if "cets-002" in text or "cets-003" in text:
            return "历史CETS编号只保留迁移说明，不作为canonical_source_id返回"
        if "gov-ind-" in text and ("政策来源" in text or "当成" in text or "来源" in text):
            return "GOV-IND治理记录不是政策来源，普通检索不返回其正文"
        if "没有章节" in text or ("没有页码" in text and "表号" in text):
            return "缺少章节、页码或表号的条款只能待补和人工复核"
        if "不在11行业" in text or "范围外行业" in text:
            return "行业不在M5冻结的11行业范围内，不套用相近行业规则"
        return None

    def _freeze_context(self, connection: sqlite3.Connection, run: dict[str, Any], index: dict[str, Any]) -> dict[str, Any]:
        row = connection.execute("SELECT * FROM assessment_run_knowledge_context WHERE assessment_run_id = ?", (run["assessment_run_id"],)).fetchone()
        if row is not None:
            return dict(row)
        created_at = str(run.get("created_at") or DECISION_DATE)
        try:
            knowledge_as_of = datetime.fromisoformat(created_at.replace("Z", "+00:00")).date().isoformat()
        except ValueError:
            knowledge_as_of = DECISION_DATE
        values = (
            run["assessment_run_id"], run["workspace_id"], run["enterprise_id"], f"run-{run['assessment_run_id']}", knowledge_as_of,
            index["index_version_id"], index["allowlist_version"], index["ranking_config_hash"], TOKENIZER_VERSION,
            INDUSTRY_MAPPING_VERSION, _sha256(_json(INDUSTRIES)), index["manifest_hash"], _now(),
        )
        connection.execute("""INSERT OR IGNORE INTO assessment_run_knowledge_context(
            assessment_run_id,workspace_id,enterprise_id,thread_id,knowledge_as_of,index_version_id,
            allowlist_version,ranking_config_hash,tokenizer_version,industry_mapping_version,
            industry_synonym_hash,source_manifest_hash,frozen_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""", values)
        return dict(connection.execute("SELECT * FROM assessment_run_knowledge_context WHERE assessment_run_id = ?", (run["assessment_run_id"],)).fetchone())

    def _source_for_member(self, connection: sqlite3.Connection, index_id: str, entry_id: str) -> sqlite3.Row | None:
        return connection.execute(
            """SELECT s.*, m.entry_type, m.chunk_id AS member_chunk_id, m.visibility AS member_visibility,
                c.chunk_type,c.title AS chunk_title,c.section_title,c.page_start,c.page_end,c.locator,c.text,
                c.verification_status AS chunk_verification_status,c.use_boundary AS chunk_use_boundary,
                c.industry_scope_json AS chunk_industry_scope,c.source_role AS chunk_source_role,
                c.official_url AS chunk_official_url,c.publisher AS chunk_publisher,c.version AS chunk_version
            FROM knowledge_index_members m
            JOIN knowledge_sources s ON s.source_uid = m.source_uid
            LEFT JOIN knowledge_chunks c ON c.chunk_id = m.chunk_id
            WHERE m.index_version_id = ? AND m.entry_id = ?""",
            (index_id, entry_id),
        ).fetchone()

    def _candidate_entries(
        self,
        connection: sqlite3.Connection,
        index_id: str,
        industry: str | None,
        roles: list[str],
        query: str,
        knowledge_as_of: str,
    ) -> list[sqlite3.Row]:
        rows = connection.execute(
            """SELECT m.entry_id,m.entry_type,m.source_uid,m.source_id,m.chunk_id,m.visibility,
                s.title,s.document_no_or_standard_no,s.source_role,s.industry_scope_json,s.official_url,
                s.publisher,s.version,s.verification_status,s.use_boundary,s.date_uncertain,s.effective_at,s.expires_at,
                c.document_id,c.chunk_type,c.title AS chunk_title,c.section_title,c.page_start,c.page_end,c.locator,c.text,
                c.normalized_search_text,c.content_hash,c.verification_status AS chunk_verification_status,
                c.use_boundary AS chunk_use_boundary,c.industry_scope_json AS chunk_industry_scope,
                c.source_role AS chunk_source_role,c.official_url AS chunk_official_url,
                c.publisher AS chunk_publisher,c.version AS chunk_version
            FROM knowledge_index_members m
            JOIN knowledge_sources s ON s.source_uid = m.source_uid
            LEFT JOIN knowledge_chunks c ON c.chunk_id = m.chunk_id
            WHERE m.index_version_id = ? AND m.visibility IN ('searchable_candidate','metadata_only')""",
            (index_id,),
        ).fetchall()
        output: list[sqlite3.Row] = []
        as_of = _parse_full_date(knowledge_as_of) or date.fromisoformat(DECISION_DATE)
        for row in rows:
            scopes = json.loads(row["chunk_industry_scope"] or row["industry_scope_json"] or "[]")
            if industry and (not scopes or ("global" not in scopes and industry not in scopes)):
                continue
            if roles and (row["chunk_source_role"] or row["source_role"]) not in roles:
                continue
            effective_at = _parse_full_date(row["effective_at"])
            expires_at = _parse_full_date(row["expires_at"])
            if effective_at and effective_at > as_of:
                continue
            if expires_at and expires_at <= as_of:
                continue
            # Metadata-only entries are a controlled exact lookup surface,
            # never ordinary full-text/keyword results.
            if row["entry_type"] == "source_metadata" and self._tier(row, query, normalize_query_tokens(query))[0] >= 2:
                continue
            output.append(row)
        return output

    @staticmethod
    def _tier(row: sqlite3.Row, query: str, tokens: list[str]) -> tuple[int, float]:
        query_norm = normalize_search_text(query)
        source_id = str(row["source_id"] or "")
        standard = str(row["document_no_or_standard_no"] or "")
        title = str(row["title"] or "")
        body = str(row["text"] or "")
        if query.strip() and (query.strip().casefold() == source_id.casefold() or query.strip().casefold() == standard.casefold()):
            return 0, 0.0
        if query_norm and query_norm == normalize_search_text(title):
            return 1, 0.0
        normalized = normalize_search_text(source_id, standard, title, row["section_title"], body)
        overlap = sum(1 for token in tokens if token in normalized)
        if overlap:
            return 2, float(-overlap)
        return 3, 0.0

    def _search_rows(self, connection: sqlite3.Connection, index: dict[str, Any], query: str, industry: str | None, roles: list[str], knowledge_as_of: str) -> list[tuple[sqlite3.Row, int, float]]:
        rows = self._candidate_entries(connection, index["index_version_id"], industry, roles, query, knowledge_as_of)
        tokens = normalize_query_tokens(query)
        candidates: list[tuple[sqlite3.Row, int, float]] = []
        if self.fts_available and tokens:
            fts_query = " OR ".join(f'"{token.replace(chr(34), "")}"' for token in tokens[:32])
            try:
                # knowledge_fts columns are entry_id, index_version_id,
                # source_id, title, standard_no, section_title, body and
                # normalized_search_text. SQLite requires one weight for every
                # column, including the three UNINDEXED columns; otherwise the
                # intended title/standard weights shift onto metadata fields.
                bm25_sql = ", ".join(f"{weight:.1f}" for weight in BM25_WEIGHTS)
                fts_rows = connection.execute(f"SELECT entry_id, bm25(knowledge_fts, {bm25_sql}) AS score FROM knowledge_fts WHERE index_version_id = ? AND knowledge_fts MATCH ? ORDER BY score, entry_id", (index["index_version_id"], fts_query)).fetchall()
                scores = {row["entry_id"]: float(row["score"]) for row in fts_rows}
            except sqlite3.OperationalError:
                scores = {}
            for row in rows:
                tier, base = self._tier(row, query, tokens)
                if row["entry_id"] in scores:
                    candidates.append((row, min(tier, 2), scores[row["entry_id"]]))
                elif tier in {0, 1}:
                    candidates.append((row, tier, base))
        else:
            for row in rows:
                tier, score = self._tier(row, query, tokens)
                if tier < 3:
                    candidates.append((row, tier, score))
        candidates.sort(key=lambda item: (
            item[1],
            source_role_priority(item[0]["chunk_source_role"] or item[0]["source_role"]),
            0 if item[0]["entry_type"] == "chunk" else 1,
            item[2],
            0 if "verified" in str(item[0]["chunk_verification_status"] or item[0]["verification_status"] or "").casefold() else 1,
            str(item[0]["source_id"] or ""),
            str(item[0]["chunk_id"] or ""),
            str(item[0]["document_id"] or ""),
        ))
        return candidates

    @staticmethod
    def _result_from_row(row: sqlite3.Row, tier: int) -> KnowledgeSearchResult:
        result_type = "chunk" if row["entry_type"] == "chunk" else "source_metadata"
        scopes = json.loads(row["chunk_industry_scope"] or row["industry_scope_json"] or "[]")
        if result_type == "chunk":
            return KnowledgeSearchResult(
                result_type="chunk", source_id=row["source_id"], document_id=row["document_id"], chunk_id=row["chunk_id"], title=row["chunk_title"] or row["title"], publisher=row["chunk_publisher"] or row["publisher"], version=row["chunk_version"] or row["version"], locator=row["locator"], excerpt=_short_text(row["text"], 1000), source_role=row["chunk_source_role"] or row["source_role"] or "other", verification_status=row["chunk_verification_status"] or row["verification_status"] or "部分核验", visibility=row["visibility"], use_boundary=row["chunk_use_boundary"] or row["use_boundary"] or "候选证据；不得自动执行", official_url=row["chunk_official_url"] or row["official_url"], industry_scope=scopes, match_tier=tier, date_uncertain=bool(row["date_uncertain"]),
            )
        return KnowledgeSearchResult(
            result_type="source_metadata", source_id=row["source_id"], title=row["title"], publisher=row["publisher"], version=row["version"], source_role=row["source_role"] or "other", verification_status=row["verification_status"] or "部分核验", visibility=row["visibility"], use_boundary=row["use_boundary"] or "正文不可用/待定位，不得引用条款", official_url=row["official_url"], industry_scope=scopes, match_tier=tier, date_uncertain=bool(row["date_uncertain"]),
        )

    def search(self, assessment_run_id: str, query: str, *, top_k: int = 5, source_roles: list[str] | None = None) -> dict[str, Any]:
        if not query or not query.strip():
            raise DomainConflictError("知识检索问题不能为空")
        if not 1 <= top_k <= 10:
            raise DomainConflictError("top_k必须在1到10之间")
        run, industry = self._run_industry(assessment_run_id)
        if industry is None:
            raise DomainConflictError("当前企业行业为空或不在11行业规范映射内，不能自动套用其他行业知识")
        # Ordinary retrieval is read-only. Index creation, replacement and
        # repair are administrator operations exposed by rebuild endpoints.
        index = self.index_status()
        if not index.get("available"):
            raise KnowledgeIndexNotReady("knowledge_index_not_ready: 本地知识索引尚未构建，请由管理员执行受控构建")
        with self.domain_store._connection() as connection:
            context = self._freeze_context(connection, run, index)
            pinned = dict(context)
            if str(pinned["index_version_id"]) != str(index["index_version_id"]):
                index_row = connection.execute("SELECT * FROM knowledge_index_versions WHERE index_version_id = ?", (pinned["index_version_id"],)).fetchone()
                if index_row is None or index_row["build_status"] != "built":
                    raise DomainConflictError("pinned_index_unavailable: 当前评估运行冻结的知识索引不可用，不能静默切换版本")
                index = {key: index_row[key] for key in index_row.keys()}
            roles = sorted(set(str(item) for item in (source_roles or []) if str(item).strip()))
            guard_reason = self._query_guard(query.strip())
            candidates = [] if guard_reason else self._search_rows(connection, index, query.strip(), industry, roles, context["knowledge_as_of"])
            selected = candidates[:top_k]
            results = [self._result_from_row(row, tier) for row, tier, _score in selected]
            warnings: list[str] = []
            if guard_reason:
                warnings.append(guard_reason)
            if not results:
                warnings.append("未找到可用证据：当前行业、知识版本或准入状态下没有可追溯结果；请补充材料或人工复核。")
            if not self.fts_available:
                warnings.append("SQLite FTS5不可用，已使用本地确定性标题/标准号/关键词降级检索。")
            if any(item.result_type == "source_metadata" for item in results):
                warnings.append("部分结果仅为来源元数据，正文不可用/待OCR/待精确定位，不得生成伪正文。")
            if any(item.date_uncertain for item in results):
                warnings.append("部分来源只有年份而没有完整生效日期，已标记date_uncertain，未按年份首日推断。")
            if not results and any(token in query for token in ("因子", "排放因子")):
                warnings.append("M5不自动调用具体排放因子值；国家因子库仅保留来源和适用性元数据。")
            returned = [item.chunk_id for item in results if item.chunk_id]
            retrieval_id = f"ret-{uuid.uuid4().hex[:16]}"
            query_hash = _sha256(query.strip())
            connection.execute(
                """INSERT INTO knowledge_retrieval_logs(
                    retrieval_id,workspace_id,assessment_run_id,enterprise_id,thread_id,query_hash,query_summary,
                    industry_filter,knowledge_as_of,index_version_id,returned_chunk_ids,created_at,retrieval_mode,fallback_reason
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (retrieval_id, run["workspace_id"], assessment_run_id, run["enterprise_id"], context["thread_id"], query_hash, self._redacted_query(query.strip()), industry, context["knowledge_as_of"], context["index_version_id"], _json(returned), _now(), "fts5" if self.fts_available else "deterministic_fallback", None if self.fts_available else "fts5_unavailable"),
            )
            payload = KnowledgeSearchResponse(
                retrieval_id=retrieval_id, assessment_run_id=assessment_run_id, workspace_id=run["workspace_id"], enterprise_id=run["enterprise_id"], enterprise_code=run["enterprise_code"], index_version_id=context["index_version_id"], allowlist_version=context["allowlist_version"], knowledge_as_of=context["knowledge_as_of"], industry_scope=[industry], query=query.strip(), results=results, warnings=warnings, degraded_mode=None if self.fts_available else "deterministic_fallback", data_notice=SIMULATED_DATA_NOTICE, untrusted_content=True,
            )
            return payload.model_dump()

    def list_retrievals(self, assessment_run_id: str) -> list[dict[str, Any]]:
        run = self.domain_store.get_assessment_run(assessment_run_id)
        with self.domain_store._connection() as connection:
            rows = connection.execute("SELECT * FROM knowledge_retrieval_logs WHERE assessment_run_id = ? ORDER BY created_at DESC, retrieval_id DESC", (assessment_run_id,)).fetchall()
            return [
                {"retrieval_id": row["retrieval_id"], "workspace_id": row["workspace_id"], "assessment_run_id": row["assessment_run_id"], "enterprise_id": row["enterprise_id"], "enterprise_code": run["enterprise_code"], "thread_id": row["thread_id"], "query_hash": row["query_hash"], "query_summary": row["query_summary"], "industry_filter": row["industry_filter"], "knowledge_as_of": row["knowledge_as_of"], "index_version_id": row["index_version_id"], "returned_chunk_ids": json.loads(row["returned_chunk_ids"] or "[]"), "created_at": row["created_at"], "retrieval_mode": row["retrieval_mode"], "fallback_reason": row["fallback_reason"],
            } for row in rows
            ]

    def get_chunk(self, assessment_run_id: str, chunk_id: str) -> dict[str, Any]:
        run, industry = self._run_industry(assessment_run_id)
        with self.domain_store._connection() as connection:
            context = connection.execute("SELECT * FROM assessment_run_knowledge_context WHERE assessment_run_id = ?", (assessment_run_id,)).fetchone()
            if context is None:
                raise DomainNotFoundError("当前运行尚未执行知识检索，无法读取证据切片")
            if context["workspace_id"] != run["workspace_id"] or context["enterprise_id"] != run["enterprise_id"]:
                raise DomainNotFoundError("证据切片不存在、未被当前运行检索返回或不属于当前行业")
            returned_ids: set[str] = set()
            for log in connection.execute(
                """SELECT returned_chunk_ids FROM knowledge_retrieval_logs
                   WHERE workspace_id = ? AND assessment_run_id = ? AND enterprise_id = ?
                     AND thread_id = ? AND index_version_id = ?""",
                (run["workspace_id"], assessment_run_id, run["enterprise_id"], context["thread_id"], context["index_version_id"]),
            ).fetchall():
                returned_ids.update(json.loads(log["returned_chunk_ids"] or "[]"))
            row = connection.execute("""SELECT c.*, s.title,s.publisher,s.version,s.official_url,s.source_role,
                s.verification_status AS source_verification_status,s.use_boundary AS source_use_boundary
                FROM knowledge_index_members m JOIN knowledge_chunks c ON c.chunk_id = m.chunk_id
                JOIN knowledge_sources s ON s.source_uid = m.source_uid
                WHERE m.index_version_id = ? AND m.chunk_id = ? AND m.visibility = 'searchable_candidate'""", (context["index_version_id"], chunk_id)).fetchone()
            scopes = json.loads(row["industry_scope_json"] or "[]") if row is not None else []
            if row is None or not industry or ("global" not in scopes and industry not in scopes) or chunk_id not in returned_ids:
                # Keep the response intentionally indistinguishable for
                # missing, cross-industry, cross-enterprise and unreturned
                # chunks; existence of another run's evidence must not leak.
                raise DomainNotFoundError("证据切片不存在、未被当前运行检索返回或不属于当前行业")
            return {
                "assessment_run_id": assessment_run_id, "workspace_id": run["workspace_id"], "enterprise_id": run["enterprise_id"], "enterprise_code": run["enterprise_code"], "thread_id": context["thread_id"], "index_version_id": context["index_version_id"], "source_id": row["source_id"], "document_id": row["document_id"], "chunk_id": row["chunk_id"], "title": row["title"], "publisher": row["publisher"], "version": row["version"], "locator": row["locator"], "excerpt": _short_text(row["text"], 4000), "source_role": row["source_role"], "verification_status": row["verification_status"], "visibility": row["visibility"], "use_boundary": row["use_boundary"], "official_url": row["official_url"], "industry_scope": scopes, "notice": "切片只作为候选证据展示，不自动改变企业事实、规则、因子、权重、阈值或授信结论。",
            }

    def list_sources(self, *, visibility: str | None = None, source_role: str | None = None) -> list[dict[str, Any]]:
        with self.domain_store._connection() as connection:
            query = "SELECT * FROM knowledge_sources WHERE 1=1"
            params: list[Any] = []
            if visibility:
                query += " AND visibility = ?"
                params.append(visibility)
            if source_role:
                query += " AND source_role = ?"
                params.append(source_role)
            rows = connection.execute(query + " ORDER BY source_id, updated_at DESC", params).fetchall()
            return [
                {"source_id": row["source_id"], "canonical_source_id": row["canonical_source_id"], "title": row["title"], "publisher": row["publisher"], "document_no_or_standard_no": row["document_no_or_standard_no"], "source_role": row["source_role"], "version": row["version"], "effective_at": row["effective_at"], "expires_at": row["expires_at"], "date_uncertain": bool(row["date_uncertain"]), "region": row["region"], "industry_scope": json.loads(row["industry_scope_json"] or "[]"), "official_url": row["official_url"], "verification_status": row["verification_status"], "admission_status": row["admission_status"], "visibility": row["visibility"], "use_boundary": row["use_boundary"], "document_id": row["document_id"], "mapping_method": row["mapping_method"],
            } for row in rows
            ]

    def _diagnostic_matches(self, connection: sqlite3.Connection, query: str, industry: str | None) -> list[sqlite3.Row]:
        tokens = normalize_query_tokens(query)
        rows = connection.execute(
            "SELECT * FROM knowledge_governance_records WHERE visibility = 'diagnostic_only' ORDER BY source_record_id"
        ).fetchall()
        matches: list[sqlite3.Row] = []
        for row in rows:
            row_industry = normalize_industry(row["industry"])
            if industry and row_industry != industry:
                continue
            searchable = normalize_search_text(row["source_record_id"], row["chunk_id"], row["industry"], row["issue_summary"], row["text"])
            if tokens and sum(token in searchable for token in tokens) >= 1:
                matches.append(row)
        return matches

    def _gold_run_contexts(self) -> list[dict[str, Any]]:
        """Return stable, explicit run contexts for Gold isolation probes."""
        contexts: list[dict[str, Any]] = []
        for workspace in sorted(self.domain_store.list_workspaces(), key=lambda item: str(item["workspace_id"])):
            runs = self.domain_store.list_assessment_runs(workspace["workspace_id"])
            for run in sorted(runs, key=lambda item: (str(item.get("enterprise_code") or ""), str(item["assessment_run_id"]))):
                contexts.append(
                    {
                        "workspace_id": run["workspace_id"],
                        "assessment_run_id": run["assessment_run_id"],
                        "enterprise_id": run["enterprise_id"],
                        "enterprise_code": run["enterprise_code"],
                        "thread_id": f"run-{run['assessment_run_id']}",
                    }
                )
        return contexts

    @staticmethod
    def _gold_probe_pairs(contexts: list[dict[str, Any]], boundary_type: str) -> list[tuple[dict[str, Any], dict[str, Any]]]:
        pairs = [
            (source, target)
            for source in contexts
            for target in contexts
            if source["assessment_run_id"] != target["assessment_run_id"]
            and source["workspace_id"] == target["workspace_id"]
            and (
                (boundary_type == "cross_enterprise" and source["enterprise_id"] != target["enterprise_id"])
                or boundary_type == "cross_run"
            )
        ]
        if boundary_type == "cross_run":
            # Prefer a same-enterprise pair when the fixture provides repeated
            # assessments, but do not depend on that shape for ordinary
            # multi-enterprise workspaces.
            pairs.sort(
                key=lambda pair: (
                    0 if pair[0]["enterprise_id"] == pair[1]["enterprise_id"] else 1,
                    str(pair[0]["enterprise_code"]),
                    str(pair[1]["enterprise_code"]),
                    str(pair[0]["assessment_run_id"]),
                    str(pair[1]["assessment_run_id"]),
                )
            )
        else:
            pairs.sort(
                key=lambda pair: (
                    str(pair[0]["enterprise_code"]),
                    str(pair[1]["enterprise_code"]),
                    str(pair[0]["assessment_run_id"]),
                    str(pair[1]["assessment_run_id"]),
                )
            )
        return pairs

    def _gold_source_chunk(self, source: dict[str, Any]) -> dict[str, Any] | None:
        """Create a source-only retrieval log with a non-global chunk."""
        try:
            _run, industry = self._run_industry(source["assessment_run_id"])
        except (DomainConflictError, DomainNotFoundError):
            return None
        if not industry:
            return None
        queries = (f"{industry} 节能 工序", "节能 工序")
        for query in queries:
            try:
                response = self.search(source["assessment_run_id"], query, top_k=10)
            except (DomainConflictError, DomainNotFoundError):
                continue
            candidates = [
                item
                for item in response["results"]
                if item.get("chunk_id")
                and industry in (item.get("industry_scope") or [])
                and "global" not in (item.get("industry_scope") or [])
            ]
            if not candidates:
                continue
            source_logs = self.list_retrievals(source["assessment_run_id"])
            source_log = next((item for item in source_logs if item["retrieval_id"] == response["retrieval_id"]), None)
            if source_log is None:
                continue
            selected = candidates[0]
            return {
                "query": query,
                "retrieval_id": response["retrieval_id"],
                "thread_id": source_log["thread_id"],
                "chunk_id": selected["chunk_id"],
                "source_id": selected["source_id"],
                "industry_scope": selected["industry_scope"],
                "returned_chunk_ids": source_log["returned_chunk_ids"],
            }
        return None

    def _run_isolation_gold_probe(self, boundary_type: str) -> dict[str, Any]:
        """Run one independently scoped enterprise/run isolation probe.

        The target run receives its own empty retrieval log, so a rejection is
        verified against the run's returned-chunk binding rather than merely
        against a missing context. Global chunks are excluded from the source
        object to avoid treating a legitimately shareable policy fragment as
        an access violation.
        """
        if boundary_type not in {"cross_enterprise", "cross_run"}:
            return {"status": "env_blocked", "boundary_type": boundary_type, "reason": "未知隔离探针边界类型"}
        contexts = self._gold_run_contexts()
        pairs = self._gold_probe_pairs(contexts, boundary_type)
        if not pairs:
            return {
                "status": "env_blocked",
                "boundary_type": boundary_type,
                "reason": "没有满足同一工作空间、运行关系和边界条件的受控运行对",
                "candidate_run_count": len(contexts),
            }

        target_query = "M5_gold_isolation_target_without_source_chunk"
        for source, target in pairs:
            source_evidence = self._gold_source_chunk(source)
            if source_evidence is None:
                continue
            try:
                target_response = self.search(target["assessment_run_id"], target_query, top_k=1)
                target_logs = self.list_retrievals(target["assessment_run_id"])
                target_log = next((item for item in target_logs if item["retrieval_id"] == target_response["retrieval_id"]), None)
                if target_log is None or source_evidence["chunk_id"] in set(target_log["returned_chunk_ids"]):
                    continue
                try:
                    self.get_chunk(target["assessment_run_id"], source_evidence["chunk_id"])
                    target_rejected = False
                    rejection_reason = None
                except DomainNotFoundError as exc:
                    target_rejected = True
                    rejection_reason = str(exc)
                checks = {
                    "same_workspace": source["workspace_id"] == target["workspace_id"],
                    "enterprise_distinct": source["enterprise_id"] != target["enterprise_id"],
                    "run_distinct": source["assessment_run_id"] != target["assessment_run_id"],
                    "thread_distinct": source_evidence["thread_id"] != target_log["thread_id"],
                    "source_log_contains_chunk": source_evidence["chunk_id"] in set(source_evidence["returned_chunk_ids"]),
                    "target_log_exists": True,
                    "target_log_excludes_chunk": source_evidence["chunk_id"] not in set(target_log["returned_chunk_ids"]),
                    "target_rejected": target_rejected,
                }
                boundary_ok = checks["enterprise_distinct"] if boundary_type == "cross_enterprise" else checks["run_distinct"]
                checks["declared_boundary_distinct"] = boundary_ok
                return {
                    "status": "passed" if all(checks.values()) else "failed",
                    "boundary_type": boundary_type,
                    "source_workspace_id": source["workspace_id"],
                    "target_workspace_id": target["workspace_id"],
                    "source_assessment_run_id": source["assessment_run_id"],
                    "target_assessment_run_id": target["assessment_run_id"],
                    "source_enterprise_id": source["enterprise_id"],
                    "target_enterprise_id": target["enterprise_id"],
                    "source_enterprise_code": source["enterprise_code"],
                    "target_enterprise_code": target["enterprise_code"],
                    "source_thread_id": source_evidence["thread_id"],
                    "target_thread_id": target_log["thread_id"],
                    "source_retrieval_id": source_evidence["retrieval_id"],
                    "target_retrieval_id": target_log["retrieval_id"],
                    "probe_query": source_evidence["query"],
                    "target_query": target_query,
                    "probe_chunk_id": source_evidence["chunk_id"],
                    "probe_source_id": source_evidence["source_id"],
                    "probe_industry_scope": source_evidence["industry_scope"],
                    "source_returned_chunk_ids": source_evidence["returned_chunk_ids"],
                    "target_returned_chunk_ids": target_log["returned_chunk_ids"],
                    "checks": checks,
                    "target_rejection_reason": rejection_reason,
                }
            except (DomainConflictError, DomainNotFoundError) as exc:
                return {
                    "status": "failed",
                    "boundary_type": boundary_type,
                    "source_assessment_run_id": source["assessment_run_id"],
                    "target_assessment_run_id": target["assessment_run_id"],
                    "reason": str(exc),
                }
        return {
            "status": "env_blocked",
            "boundary_type": boundary_type,
            "reason": "候选运行对中没有找到仅记录于源运行日志、且不属于global范围的可用切片",
            "candidate_run_count": len(contexts),
        }

    def run_gold_tests(self) -> list[dict[str, Any]]:
        tests = _load_rows(self.assets.gold_tests, "01_金标准")
        index = self.index_status()

        def split_ids(value: Any) -> list[str]:
            return [item.strip() for item in re.split(r"[,，;；\s]+", str(value or "")) if item.strip()]

        if not index.get("available"):
            return [
                {
                    "test_id": str(row.get("test_id") or ""),
                    "status": "not_executed",
                    "passed": False,
                    "query": str(row.get("query") or ""),
                    "reason": "knowledge_index_not_ready: Gold必须在管理员构建成功的冻结索引上运行",
                    "notice": "这是内部检索门槛测试结果，不是正式检索准确率或真实业务效果。",
                }
                for row in tests
            ]

        isolation_probes: dict[str, dict[str, Any]] = {}
        with self.domain_store._connection() as connection:
            copper_scope_rows = self._search_rows(connection, index, "钢铁 工序", "冶金行业铜", [], DECISION_DATE)
            unscoped_rows = self._search_rows(connection, index, "钢铁 工序", None, [], DECISION_DATE)
            unscoped_steel_ids = {str(item[0]["source_id"]) for item in unscoped_rows}
            copper_steel_ids = {str(item[0]["source_id"]) for item in copper_scope_rows}
            leaked_diagnostic_members = connection.execute(
                "SELECT COUNT(*) AS count FROM knowledge_index_members WHERE index_version_id = ? AND visibility = 'diagnostic_only'",
                (index["index_version_id"],),
            ).fetchone()["count"]
            mutation_probe = {
                "status": "passed" if (unscoped_steel_ids & {"STD-007", "CETS-VG-001"}) and not (copper_steel_ids & {"STD-007", "CETS-VG-001"}) and leaked_diagnostic_members == 0 else "failed",
                "unscoped_steel_ids": sorted(unscoped_steel_ids & {"STD-007", "CETS-VG-001"}),
                "copper_scope_steel_ids": sorted(copper_steel_ids & {"STD-007", "CETS-VG-001"}),
                "diagnostic_members": leaked_diagnostic_members,
            }
            governance_rows = connection.execute("SELECT * FROM knowledge_governance_records").fetchall()
            governance_lookup: dict[str, dict[str, Any]] = {}
            for item in governance_rows:
                record = dict(item)
                for key in ("governance_record_id", "source_record_id", "chunk_id"):
                    if record.get(key):
                        governance_lookup[str(record[key])] = record
            # A metadata row is not a usable Gold target. Only a searchable
            # index member or an explicitly governed diagnostic record counts
            # as an available target; blocked/metadata-only registrations must
            # still exercise the correct-degrade path.
            searchable_source_ids = {
                str(row["source_id"])
                for row in connection.execute(
                    "SELECT DISTINCT source_id FROM knowledge_index_members WHERE index_version_id = ? AND visibility = 'searchable_candidate'",
                    (index["index_version_id"],),
                ).fetchall()
                if row["source_id"]
            }
            searchable_chunk_ids = {
                str(row["chunk_id"])
                for row in connection.execute(
                    "SELECT DISTINCT chunk_id FROM knowledge_index_members WHERE index_version_id = ? AND visibility = 'searchable_candidate' AND chunk_id IS NOT NULL",
                    (index["index_version_id"],),
                ).fetchall()
                if row["chunk_id"]
            }

            results: list[dict[str, Any]] = []
            for row in tests:
                test_id = str(row.get("test_id") or "")
                query = str(row.get("query") or "").strip()
                industry_value = row.get("industry")
                industry = None if industry_value in (None, "", "通用") else normalize_industry(industry_value)
                expected_source_ids = split_ids(row.get("expected_source_ids"))
                expected_chunk_ids = split_ids(row.get("expected_chunk_ids"))
                forbidden_ids = split_ids(row.get("forbidden_source_ids"))
                expected_behavior = str(row.get("expected_behavior") or "")
                top_k = int(row.get("top_k") or 5)
                guard_reason = self._query_guard(query)
                public_rows = [] if guard_reason else self._search_rows(connection, index, query, industry, [], DECISION_DATE)[:top_k]
                actual_top_k = [
                    {
                        "source_id": item[0]["source_id"],
                        "chunk_id": item[0]["chunk_id"],
                        "document_id": item[0]["document_id"],
                        "title": item[0]["chunk_title"] or item[0]["title"],
                        "visibility": item[0]["visibility"],
                        "source_role": item[0]["chunk_source_role"] or item[0]["source_role"],
                        "verification_status": item[0]["chunk_verification_status"] or item[0]["verification_status"],
                        "use_boundary": item[0]["chunk_use_boundary"] or item[0]["use_boundary"],
                        "industry_scope": json.loads(item[0]["chunk_industry_scope"] or item[0]["industry_scope_json"] or "[]"),
                        "match_tier": item[1],
                        "score": item[2],
                    }
                    for item in public_rows
                ]
                actual_source_ids = sorted({str(item["source_id"]) for item in actual_top_k if item.get("source_id")})
                actual_chunk_ids = sorted({str(item["chunk_id"]) for item in actual_top_k if item.get("chunk_id")})
                diagnostic_rows = self._diagnostic_matches(connection, query, industry)
                diagnostic_ids = sorted({str(item[key]) for item in diagnostic_rows for key in ("source_record_id", "chunk_id") if item[key]})
                diagnostic_expected = [item for item in expected_source_ids + expected_chunk_ids if item.startswith(("PATH-", "GOV-IND-", "CHUNK-PATH-", "CHUNK-IND-"))]
                missing_diagnostic = [item for item in diagnostic_expected if item not in diagnostic_ids]
                expected_ids = expected_source_ids + expected_chunk_ids
                target_available_ids = sorted({item for item in expected_ids if item in searchable_source_ids or item in searchable_chunk_ids or item in governance_lookup})
                target_missing_ids = sorted(set(expected_ids) - set(diagnostic_ids) - set(actual_source_ids) - set(actual_chunk_ids))
                target_hit = not target_missing_ids
                fallback_allowed = any(marker in expected_behavior for marker in ("若无可用证据", "不返回权威结论", "触发补问", "人工复核"))
                forbidden_hit = [item for item in forbidden_ids if item in actual_source_ids or item in actual_chunk_ids]
                actual_authoritative = [item for item in actual_top_k if item["visibility"] == "searchable_candidate" and item["source_role"] in {"official_standard", "official_policy", "official_methodology", "regulatory_guidance"}]
                no_authoritative_result = not actual_authoritative
                returned_results_compliant = all(
                    item["visibility"] in {"searchable_candidate", "metadata_only"}
                    and (not industry or "global" in item["industry_scope"] or industry in item["industry_scope"])
                    and "自动执行" in str(item.get("use_boundary") or "")
                    for item in actual_top_k
                )
                warnings = []
                if guard_reason:
                    warnings.append(guard_reason)
                if no_authoritative_result:
                    warnings.append("未返回可作为正式权威结论的结果；按金标准降级为补充材料/人工复核。")
                if not actual_top_k:
                    warnings.append("实际top-k为空")
                if not returned_results_compliant:
                    warnings.append("实际返回结果未同时满足行业范围、可见性和候选证据使用边界")

                if test_id in {"N-S2-05", "N-S2-06"}:
                    boundary_type = "cross_enterprise" if test_id == "N-S2-05" else "cross_run"
                    isolation_probe = isolation_probes.get(boundary_type)
                    if isolation_probe is None:
                        isolation_probe = self._run_isolation_gold_probe(boundary_type)
                        isolation_probes[boundary_type] = isolation_probe
                    passed = isolation_probe.get("status") == "passed"
                    status = "passed" if passed else str(isolation_probe.get("status") or "failed")
                    reason = (
                        f"已独立执行{boundary_type}隔离探针，源运行切片只存在于源运行检索日志，目标运行读取被拒绝"
                        if passed
                        else str(isolation_probe.get("reason") or "真实运行隔离探针未通过")
                    )
                    mode = "real_run_isolation_probe"
                elif test_id.startswith("N-S2-"):
                    passed = not forbidden_hit and guard_reason is not None and not actual_authoritative and (test_id != "N-S2-01" or mutation_probe["status"] == "passed")
                    status = "passed" if passed else "failed"
                    reason = guard_reason or ("未阻断权威结果" if actual_authoritative else "未命中禁止对象，但未触发预期负向边界" if not forbidden_hit else f"命中禁止对象：{forbidden_hit}")
                    mode = "real_service_negative_search"
                else:
                    passed = not forbidden_hit and not missing_diagnostic and returned_results_compliant
                    if missing_diagnostic:
                        status = "failed"
                        reason = f"金标准要求的治理对象未在diagnostic_only仓储中找到：{missing_diagnostic}"
                    elif diagnostic_expected:
                        status = "passed"
                        reason = "已通过实际治理仓储查询定位对象；对象保持diagnostic_only，普通检索结果不返回其正文"
                    elif expected_ids and not target_hit:
                        if not target_available_ids and fallback_allowed:
                            status = "correct_degrade"
                            passed = False
                            warnings.append(f"金标准期望来源/切片未纳入当前治理资产：{target_missing_ids}；未将其他结果冒充目标，已保留补问/人工复核边界。")
                            reason = "期望来源/切片当前不可用，且金标准明确允许降级；其他返回结果仍满足行业、可见性和候选证据边界，未命中目标不计入真正通过"
                        else:
                            status = "failed"
                            passed = False
                            reason = f"金标准期望来源/切片未命中：{target_missing_ids}"
                    elif no_authoritative_result:
                        status = "correct_degrade"
                        reason = "无权威正文结果且已生成证据不足/人工复核警告，符合降级预期"
                    else:
                        status = "passed"
                        reason = "普通检索返回当前行业可见候选证据；行业、可见性和使用边界检查通过，未将结果升级为正式规则"
                    if forbidden_hit:
                        status = "failed"
                    mode = "deterministic_ordinary_and_governance_check"
                results.append(
                    {
                        "test_id": test_id,
                        "module_id": row.get("module_id"),
                        "context": row.get("context"),
                        "industry": industry or industry_value,
                        "query": query,
                        "expected_source_ids": expected_source_ids,
                        "expected_chunk_ids": expected_chunk_ids,
                        "expected_behavior": expected_behavior,
                        "forbidden_source_ids": forbidden_ids,
                        "actual_top_k": actual_top_k,
                        "actual_source_ids": actual_source_ids,
                        "actual_chunk_ids": actual_chunk_ids,
                        "diagnostic_matches": diagnostic_ids,
                        "target_expected": expected_ids,
                        "target_available": target_available_ids,
                        "target_missing": target_missing_ids,
                        "target_hit": target_hit,
                        "fallback_allowed": fallback_allowed,
                        "visibility": sorted({str(item.get("visibility")) for item in actual_top_k}),
                        "roles": sorted({str(item.get("source_role")) for item in actual_top_k}),
                        "degrade_warnings": warnings,
                        "no_authoritative_result": no_authoritative_result,
                        "forbidden_hit": forbidden_hit,
                        "returned_count": len(actual_top_k),
                        "returned_results_compliant": returned_results_compliant,
                        "guard_reason": guard_reason,
                        "passed": passed,
                        "status": status,
                        "reason": reason,
                        "mode": mode,
                        "mutation_resistance": mutation_probe,
                        "isolation_probe": isolation_probe if test_id in {"N-S2-05", "N-S2-06"} else None,
                        "notice": "这是内部检索门槛测试结果，不是正式检索准确率或真实业务效果。",
                    }
                )
            return results
