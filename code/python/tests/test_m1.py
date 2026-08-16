from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app import main
from app.config import MAX_UPLOAD_BYTES, REQUIRED_HEADERS
from app.m1_core import deterministic_projection
from app.store import BatchStore


def make_workbook(
    *,
    missing_reference: bool = False,
    duplicate_energy_key: bool = False,
    bad_header: bool = False,
    empty_key: bool = False,
    orphan_key: bool = False,
    production_unit_mismatch: bool = False,
    negative_value: bool = False,
    bad_ratio: bool = False,
    bad_boolean: bool = False,
    invalid_numeric: bool = False,
    empty_subindustry: bool = False,
) -> bytes:
    """Create a tiny labelled simulation fixture; it is not competition data."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers in REQUIRED_HEADERS.items():
        if missing_reference and sheet_name == "转型规划结论":
            continue
        worksheet = workbook.create_sheet(sheet_name)
        actual_headers = list(headers)
        if bad_header and sheet_name == "基本信息":
            actual_headers[-1] = "成立年份（错误表头）"
        worksheet.append(actual_headers)

    basic = workbook["基本信息"]
    basic.append(["" if empty_key else "TFTEST01", "钢铁", None if empty_subindustry else "钢铁", "江西", "中型", 2010])
    basic.append(["TFTEST02", "钢铁", "钢铁", "江西", "小型", 2018])

    energy = workbook["能耗信息"]
    first_energy = [
        "TFTEST01", "电力、天然气", 2,
        100, 120, None, None, None, None, None, None, None, None, None, None, None, None,
        10, 20, None if production_unit_mismatch else "吨", "bad" if invalid_numeric else 1000, 1100,
    ]
    if negative_value:
        first_energy[3] = -1
    energy.append(first_energy)
    second_energy = [
        "TFTEST02", "电力", 1,
        30, None, None, None, None, None, None, None, None, None, None, None, None, None,
        None, None, None, 200, 210,
    ]
    energy.append(second_energy)
    if duplicate_energy_key:
        energy.append(list(energy.iter_rows(min_row=2, max_row=2, values_only=True))[0])
    if orphan_key:
        orphan = list(second_energy)
        orphan[0] = "TFORPHAN"
        energy.append(orphan)

    supplement = workbook["补充信息"]
    supplement.append(["TFTEST01", None if empty_subindustry else "钢铁", "电炉", 30, "maybe" if bad_boolean else "是", "否", "否", 1.5 if bad_ratio else 0.2, "能耗偏高", "希望节能改造"])
    supplement.append(["TFTEST02", "钢铁", "轧机", 2, "否", "否", "否", 0.0, "数据不足", "希望补齐数据"])

    if not missing_reference:
        reference = workbook["转型规划结论"]
        reference.append(["TFTEST01", "参考用能", "参考关联", "参考改进", "参考路径", "近阶段参考", "中期参考", "长期参考", "参考要点"])
        reference.append(["TFTEST02", "参考用能2", "参考关联2", "参考改进2", "参考路径2", "近阶段参考2", "中期参考2", "长期参考2", "参考要点2"])

    catalog = workbook["转型目录"]
    catalog.append(["钢铁", "钢铁", "电炉节能改造", "降低能源消耗"])
    catalog.append(["钢铁", None, "行业级能效提升", "行业级候选"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@pytest.fixture()
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> TestClient:
    monkeypatch.setattr(main, "store", BatchStore(tmp_path / "runtime"))
    return TestClient(main.app)


def upload(client: TestClient, content: bytes) -> dict:
    response = client.post(
        "/api/v1/documents",
        files={"file": ("m1_simulated.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200, response.text
    return response.json()


def test_upload_validates_five_tables_and_registers_original(client: TestClient) -> None:
    summary = upload(client, make_workbook())
    assert summary["status"] == "passed"
    assert summary["simulated_data"] is True
    assert summary["validation"]["actual_sheets"] == list(REQUIRED_HEADERS)
    assert summary["validation"]["company_codes"] == ["TFTEST01", "TFTEST02"]
    assert summary["validation"]["field_contract_version"] == "workbook-contract-v0.1"
    assert summary["quality_overview"]["enterprise_derived"]["issue_count"] > 0
    assert summary["quality_overview"]["enterprise_derived"]["affected_company_count"] > 0
    source_path = Path(main.store.source_path(summary["batch_id"]))
    assert source_path.exists()
    assert source_path.stat().st_mode & 0o777 == 0o444


def test_company_detail_energy_missing_catalog_and_reference_isolation(client: TestClient) -> None:
    summary = upload(client, make_workbook())
    batch_id = summary["batch_id"]
    detail_response = client.get(f"/api/v1/companies/TFTEST01?batch_id={batch_id}")
    assert detail_response.status_code == 200, detail_response.text
    detail = detail_response.json()
    assert "建议改进方向" not in detail["input_data"]
    assert detail["reference_comparison"]["leakage_guard"]["status"] == "passed"
    assert any(item["status"] == "成对缺失" for item in detail["energy_trend"]["resources"])
    assert any(item["rule"] == "annual_pair_missing" for item in detail["quality_issues"])
    assert any(item["rule"] == "equipment_age_warning" for item in detail["quality_issues"])
    assert detail["catalog_matches"]["candidates"][0]["catalog_row_id"].startswith("转型目录!A")
    assert "score" not in detail["catalog_matches"]["candidates"][0]
    assert detail["catalog_matches"]["candidates"][0]["provisional_rule"] is True
    assert detail["boundaries"]["formal_carbon_calculation"] is False


def test_reference_comparison_stays_available_without_becoming_input(client: TestClient) -> None:
    summary = upload(client, make_workbook())
    detail = client.get(f"/api/v1/companies/TFTEST02?batch_id={summary['batch_id']}").json()
    reference = detail["reference_comparison"]
    assert reference["reference_fields_present"] == 8
    assert len(reference["comparison_items"]) == 5
    assert any(item["status"] == "未发现字面重合" for item in reference["comparison_items"])
    assert reference["leakage_guard"]["reference_fields_excluded_from_input"]
    assert all(field not in detail["input_data"] for field in reference["leakage_guard"]["reference_fields_excluded_from_input"])


def test_invalid_structure_is_readable_and_not_silently_skipped(client: TestClient) -> None:
    summary = upload(client, make_workbook(missing_reference=True, bad_header=True))
    assert summary["status"] == "failed"
    issues = summary["validation"]["validation_issues"]
    assert any(issue["rule"] == "required_sheet" for issue in issues)
    assert any(issue["rule"] == "required_headers" for issue in issues)


def test_duplicate_key_is_located(client: TestClient) -> None:
    summary = upload(client, make_workbook(duplicate_energy_key=True))
    assert summary["status"] == "failed"
    assert any(issue["rule"] == "unique_key" and issue["sheet_name"] == "能耗信息" for issue in summary["validation"]["validation_issues"])


@pytest.mark.parametrize(
    ("option", "rule"),
    [
        ("empty_key", "non_empty_key"),
        ("orphan_key", "cross_sheet_key_set"),
        ("production_unit_mismatch", "production_unit_pair_mismatch"),
        ("negative_value", "non_negative"),
        ("bad_ratio", "ratio_range"),
        ("bad_boolean", "boolean_enum"),
    ],
)
def test_negative_quality_cases_are_located(client: TestClient, option: str, rule: str) -> None:
    summary = upload(client, make_workbook(**{option: True}))
    if rule == "production_unit_pair_mismatch":
        assert summary["status"] == "passed"
        assert summary["quality_overview"]["enterprise_derived"]["by_rule"][rule] == 1
    else:
        assert summary["status"] == "failed"
        assert any(issue["rule"] == rule for issue in summary["validation"]["validation_issues"])


def test_invalid_numeric_batch_blocks_follow_up_instead_of_crashing(client: TestClient) -> None:
    summary = upload(client, make_workbook(invalid_numeric=True))
    assert summary["status"] == "failed"
    assert summary["quality_overview"]["enterprise_derived"]["status"] == "blocked"
    companies = client.get(f"/api/v1/batches/{summary['batch_id']}/companies").json()
    assert companies["companies"] == []
    assert companies["follow_up_allowed"] is False
    detail = client.get(f"/api/v1/companies/TFTEST01?batch_id={summary['batch_id']}")
    assert detail.status_code == 409
    report = client.post("/api/v1/reports/basic", json={"batch_id": summary["batch_id"], "company_code": "TFTEST01"})
    assert report.status_code == 409


def test_malformed_empty_and_non_xlsx_uploads_are_readable(client: TestClient) -> None:
    malformed = client.post(
        "/api/v1/documents",
        files={"file": ("damaged.xlsx", b"not-an-xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert malformed.status_code == 200
    failed = malformed.json()
    assert failed["status"] == "failed"
    assert client.get(f"/api/v1/companies/TFTEST01?batch_id={failed['batch_id']}").status_code == 409

    empty = client.post("/api/v1/documents", files={"file": ("empty.xlsx", b"", "application/octet-stream")})
    assert empty.status_code == 400
    not_xlsx = client.post("/api/v1/documents", files={"file": ("data.csv", b"a,b", "text/csv")})
    assert not_xlsx.status_code == 400


def test_upload_size_limit_and_artifact_path_boundary(client: TestClient) -> None:
    oversized = client.post(
        "/api/v1/documents",
        files={"file": ("oversized.xlsx", b"x" * (MAX_UPLOAD_BYTES + 1), "application/octet-stream")},
    )
    assert oversized.status_code == 413
    summary = upload(client, make_workbook())
    report = client.post("/api/v1/reports/basic", json={"batch_id": summary["batch_id"], "company_code": "TFTEST01"})
    assert report.status_code == 200
    assert not Path(report.json()["report_relative_path"]).is_absolute()
    with pytest.raises(ValueError):
        main.store.load_json(summary["batch_id"], "../metadata.json")


def test_empty_catalog_detail_requires_manual_review(client: TestClient) -> None:
    summary = upload(client, make_workbook(empty_subindustry=True))
    detail = client.get(f"/api/v1/companies/TFTEST01?batch_id={summary['batch_id']}").json()
    catalog = detail["catalog_matches"]
    assert catalog["status"] == "缺少细分行业/产品信息"
    assert catalog["manual_review_required"] is True
    assert all("score" not in item for item in catalog["candidates"])


def test_basic_report_has_boundary_statement_and_is_repeatable(client: TestClient) -> None:
    first = upload(client, make_workbook())
    second = upload(client, make_workbook())
    first_detail = client.get(f"/api/v1/companies/TFTEST01?batch_id={first['batch_id']}").json()
    second_detail = client.get(f"/api/v1/companies/TFTEST01?batch_id={second['batch_id']}").json()
    assert deterministic_projection(first_detail) == deterministic_projection(second_detail)
    first_report = client.post("/api/v1/reports/basic", json={"batch_id": first["batch_id"], "company_code": "TFTEST01"})
    assert first_report.status_code == 200, first_report.text
    body = first_report.json()
    assert "模拟数据" in body["markdown"]
    assert "最终评分" in body["markdown"]
    assert "参考结论" in body["markdown"]
    assert "暂定排序值（非评分）" in body["markdown"]
    assert "明确语义映射" in body["markdown"]
    recovered = client.get(f"/api/v1/reports/{first['batch_id']}/{body['report_id']}")
    assert recovered.status_code == 200
    assert recovered.json()["analysis"]["company_code"] == "TFTEST01"
